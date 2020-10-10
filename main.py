# -*-coding: utf-8 -
import sys, os, re, datetime, copy, json 
import logging
import openpyxl
import resource_rc
import util, kw_util
import user_setting

from PyQt5 import QtCore, QtWidgets
from PyQt5.QtCore import QObject, pyqtSlot, pyqtSignal, QUrl, QEvent
from PyQt5.QtCore import QStateMachine, QState, QTimer, QFinalState
from PyQt5.QtWidgets import QApplication
from PyQt5.QAxContainer import QAxWidget
from mainwindow_ui import Ui_MainWindow


from oauth2client.service_account import ServiceAccountCredentials
import gspread

scope = ['https://spreadsheets.google.com/feeds',
         'https://www.googleapis.com/auth/drive']
credentials = ServiceAccountCredentials.from_json_keyfile_name(
        user_setting.GOOGLE_SPREAD_AUTH_JSON_FILE, scope)
gc = gspread.authorize(credentials) 



TRADING_INFO_GETTING_TIME = [15, 45] # 트레이딩 정보를 저장하기 시작하는 시간
TR_TIME_LIMIT_MS = 3800 # 키움 증권에서 정의한 연속 TR 시 필요 딜레이 

INTERESTED_STOCKS_FILE_PATH = "log" + os.path.sep +  "interested_stocks.json"
CHEGYEOL_INFO_FILE_PATH = "log" + os.path.sep +  "chegyeol.json"
JANGO_INFO_FILE_PATH =  "log" + os.path.sep + "jango.json"
CHEGYEOL_INFO_EXCEL_FILE_PATH = "log" + os.path.sep +  "chegyeol.xlsx" 

###################################################################################################
###################################################################################################
class CloseEventEater(QObject):
    def eventFilter(self, obj, event):
        if( event.type() == QEvent.Close):
            test_make_jangoInfo()
            return True
        else:
            return super(CloseEventEater, self).eventFilter(obj, event)

class KiwoomConditon(QObject):
    sigInitOk = pyqtSignal()
    sigConnected = pyqtSignal()
    sigDisconnected = pyqtSignal()
    sigTryConnect = pyqtSignal()
    sigGetConditionCplt = pyqtSignal()
    sigSelectCondition = pyqtSignal()
    sigReselectCondition = pyqtSignal()

    sigStateStop = pyqtSignal()
    sigStockComplete = pyqtSignal()

    sigConditionOccur = pyqtSignal()
    sigRequestMinuteCandleInfo = pyqtSignal()
    sigRequestEtcInfo = pyqtSignal()

    sigGetBasicInfo = pyqtSignal()
    sigDetermineBuy= pyqtSignal()
    sigGetHogaInfo = pyqtSignal()
    sigTrWaitComplete = pyqtSignal()

    sigWaitTr = pyqtSignal()
    sigNoWaitTr = pyqtSignal()
    sigRequestRealHogaComplete = pyqtSignal()
    sigError = pyqtSignal()
    sigRequestJangoComplete = pyqtSignal()
    sigCalculateStoplossComplete = pyqtSignal()
    sigStartProcessBuy = pyqtSignal()
    sigStopProcessBuy = pyqtSignal()
    sigTerminating = pyqtSignal()

    sigRealInfoArrived = pyqtSignal(str, str, list)

    def __init__(self):
        super().__init__()
        self.ocx = QAxWidget("KHOPENAPI.KHOpenAPICtrl.1")
        self.fsm = QStateMachine()
        self.account_list = []
        self.timerSystem = QTimer()
        self.lineCmdText = ''

        self.maesuProhibitCodeList = [] # 종목 거래 금지 list 

        self.upjongInfo = {'코스피': {}, '코스닥': {} } # { 'yupjong_code': { '현재가': 222, ...} }
        self.michegyeolInfo = {}
        self.jangoInfo = {} # { 'jongmok_code': { '이익실현가': 222, ...}}
        self.jangoInfoFromFile = {} # TR 잔고 정보 요청 조회로는 얻을 수 없는 데이터를 파일로 저장하고 첫 실행시 로드함  
        self.chegyeolInfo = {} # { '날짜' : [ [ '주문구분', '매도', '분할매수이력', '체결가' , '체결수량', '미체결수량'] ] }
        self.conditionOccurList = [] # 조건 진입이 발생한 모든 리스트 저장하고 매수 결정에 사용되는 모든 정보를 저장함  [ {'종목코드': code, ...}] 
        self.conditionRemoveList = [] # 조건 이탈이 발생한 모든 리스트 저장 
        self.conditionStoplossList = {'1분': [], '3분': [], '15분': [], '30분': [] }# 기존 조건 진입후 손절 조건을 판단하기 위함 

        self.kospiCodeList = () 
        self.kosdaqCodeList = () 

        self.createState()
        self.createConnection()
        self.currentTime = datetime.datetime.now()
        self.current_condition_name = ''

        self.kospi_updown = 0 
        self.kosdaq_updown = 0 

        # 잔고 정보 저장시 저장 제외될 키 값들 
        self.jango_remove_keys = [ 
            '매도호가1', '매도호가2', '매도호가3', '매도호가수량1', '매도호가수량2', '매도호가수량3','매도호가총잔량',
            '매수호가1', '매수호가2', '매수호가3', '매수호가수량1', '매수호가수량2', '매수호가수량3', '매수호가총잔량',
            '현재가', '호가시간', '세금', '전일종가', '현재가', '종목번호', '수익율', '수익', '잔고' , '매도중', '시가', '고가', '저가', '장구분', 
            '거래량', '등락율', '전일대비', '기준가', '상한가', '하한가',
            '일{}봉'.format(user_setting.MAX_SAVE_CANDLE_COUNT), '{}분{}봉'.format(user_setting.REQUEST_MINUTE_CANDLE_TYPE, user_setting.MAX_SAVE_CANDLE_COUNT)  ]
        
    def createState(self):
        # state defintion
        mainState = QState(self.fsm)       
        stockCompleteState = QState(self.fsm)
        finalState = QFinalState(self.fsm)
        self.fsm.setInitialState(mainState)
        
        initState = QState(mainState)
        disconnectedState = QState(mainState)
        connectedState = QState(QtCore.QState.ParallelStates, mainState)

        systemState = QState(connectedState)
        
        initSystemState = QState(systemState)
        waitingTradeSystemState = QState(systemState)
        standbySystemState = QState(systemState)
        requestingJangoSystemState = QState(systemState)
        terminatingSystemState = QState(systemState)
        
        #transition defition
        mainState.setInitialState(initState)
        mainState.addTransition(self.sigStateStop, finalState)
        mainState.addTransition(self.sigStockComplete, stockCompleteState)
        stockCompleteState.addTransition(self.sigStateStop, finalState)
        initState.addTransition(self.sigInitOk, disconnectedState)
        disconnectedState.addTransition(self.sigConnected, connectedState)
        disconnectedState.addTransition(self.sigTryConnect, disconnectedState)
        connectedState.addTransition(self.sigDisconnected, disconnectedState)
        
        systemState.setInitialState(initSystemState)
        initSystemState.addTransition(self.sigGetConditionCplt, requestingJangoSystemState)
        requestingJangoSystemState.addTransition(self.sigRequestJangoComplete, waitingTradeSystemState)

        waitingTradeSystemState.addTransition(self.sigSelectCondition, standbySystemState)

        standbySystemState.addTransition(self.sigTerminating,  terminatingSystemState )
        standbySystemState.addTransition(self.sigReselectCondition, waitingTradeSystemState )
        
        #state entered slot connect
        mainState.entered.connect(self.mainStateEntered)
        stockCompleteState.entered.connect(self.stockCompleteStateEntered)
        initState.entered.connect(self.initStateEntered)
        disconnectedState.entered.connect(self.disconnectedStateEntered)
        connectedState.entered.connect(self.connectedStateEntered)
        
        systemState.entered.connect(self.systemStateEntered)
        initSystemState.entered.connect(self.initSystemStateEntered)
        waitingTradeSystemState.entered.connect(self.waitingTradeSystemStateEntered)
        requestingJangoSystemState.entered.connect(self.requestingJangoSystemStateEntered)
        standbySystemState.entered.connect(self.standbySystemStateEntered)
        terminatingSystemState.entered.connect(self.terminatingSystemStateEntered)
    
        # processBuy definition
        processBuyState = QState(connectedState)
        initProcessBuyState = QState(processBuyState)
        standbyProcessBuyState = QState(processBuyState)
        requestEtcInfoProcessBuyState = QState(processBuyState)
        determineBuyProcessBuyState = QState(processBuyState)
        waitingTRlimitProcessBuyState = QState(processBuyState)

        processBuyState.setInitialState(initProcessBuyState)
        initProcessBuyState.addTransition(self.sigStartProcessBuy, standbyProcessBuyState)

        standbyProcessBuyState.addTransition(self.sigConditionOccur, standbyProcessBuyState)
        standbyProcessBuyState.addTransition(self.sigRequestEtcInfo, requestEtcInfoProcessBuyState)
        standbyProcessBuyState.addTransition(self.sigStopProcessBuy, initProcessBuyState)

        requestEtcInfoProcessBuyState.addTransition(self.sigWaitTr, waitingTRlimitProcessBuyState )
        requestEtcInfoProcessBuyState.addTransition(self.sigDetermineBuy, determineBuyProcessBuyState)
        requestEtcInfoProcessBuyState.addTransition(self.sigError, standbyProcessBuyState )

        determineBuyProcessBuyState.addTransition(self.sigNoWaitTr, standbyProcessBuyState)
        determineBuyProcessBuyState.addTransition(self.sigWaitTr, waitingTRlimitProcessBuyState)

        waitingTRlimitProcessBuyState.addTransition(self.sigTrWaitComplete, standbyProcessBuyState)

        processBuyState.entered.connect(self.processBuyStateEntered)
        initProcessBuyState.entered.connect(self.initProcessBuyStateEntered)
        standbyProcessBuyState.entered.connect(self.standbyProcessBuyStateEntered)
        requestEtcInfoProcessBuyState.entered.connect(self.requestEtcInfoProcessBuyStateEntered)
        determineBuyProcessBuyState.entered.connect(self.determineBuyProcessBuyStateEntered)
        waitingTRlimitProcessBuyState.entered.connect(self.waitingTRlimitProcessBuyStateEntered)
                
        #fsm start
        finalState.entered.connect(self.finalStateEntered)
        self.fsm.start()

        pass

    def initQmlEngine(self):
        self.qmlEngine.load(QUrl('qrc:///qml/main.qml'))        
        self.rootObject = self.qmlEngine.rootObjects()[0]
        self.rootObject.startClicked.connect(self.onStartClicked)
        self.rootObject.restartClicked.connect(self.onRestartClicked)
        self.rootObject.requestJangoClicked.connect(self.onRequestJangoClicked)
        self.rootObject.chegyeolClicked.connect(self.onChegyeolClicked)
        self.rootObject.testClicked.connect(self.onTestClicked)

        rootContext = self.qmlEngine.rootContext()
        rootContext.setContextProperty("model", self)
        pass
    @pyqtSlot()
    def onBtnMakeExcelClicked(self):
        print(util.whoami())
        self.make_excel(CHEGYEOL_INFO_EXCEL_FILE_PATH, self.chegyeolInfo)
        pass
    @pyqtSlot()
    def onBtnStartClicked(self):
        print(util.whoami())
        self.sigInitOk.emit()
        
    @pyqtSlot()
    def onBtnJangoClicked(self):
        self.printStockInfo()

    @pyqtSlot()
    def onBtnYupjongClicked(self):
        self.printYupjongInfo()
        pass

    @pyqtSlot()
    def onBtnChegyeolClicked(self):
        self.printChegyeolInfo('all')

    @pyqtSlot()
    def onBtnRunClicked(self):
        arg = self.lineCmdText
        # if( arg ):
        #     eval(arg)
        self.sigRealInfoArrived.emit("124", "test", [1,2,3])
        pass

    @pyqtSlot()
    def onBtnConditionClicked(self):
        items = self.getCodeListConditionOccurList()
        buffer = [] 

        for item in items: 
            boyou = "보유" if item in self.jangoInfo else ""
            log_string = "{} {} ({}-{})".format(
                item,  
                self.getMasterCodeName(item), 
                self.getMasterStockInfo(item),
                boyou) 

            if( boyou == "보유" ):
                buffer.append(log_string)
            else:
                buffer.insert(0, log_string)
        
        print( '\n'.join(buffer) )
        pass
    
    @pyqtSlot(str)
    def onLineCmdTextChanged(self, str):
        self.lineCmdText = str
        pass

    @pyqtSlot(str)
    def onTestClicked(self, arg):
        print(util.whoami() + ' ' + arg)
        # eval(arg)
        pass


    @pyqtSlot(str, str, list)
    def onRealInfoArrived(self, jongmok_code, real_data_type, result_list ):
        # print("{} {} {}".format(self.getMasterCodeName(jongmok_code),  real_data_type, result_list))

        #잔고 실시간 정보 upate
        if( jongmok_code in self.jangoInfo ):
            self.setRealData(real_data_type, self.jangoInfo[jongmok_code], result_list)
        #조건 발생 종목 정보 update
        for item_dict in self.conditionOccurList:
            if( item_dict['종목코드'] == jongmok_code ):
                self.setRealData(real_data_type, item_dict, result_list)
        pass

    def createConnection(self):
        self.ocx.OnEventConnect[int].connect(self._OnEventConnect)
        self.ocx.OnReceiveMsg[str, str, str, str].connect(self._OnReceiveMsg)
        self.ocx.OnReceiveTrData[str, str, str, str, str,
                                    int, str, str, str].connect(self._OnReceiveTrData)
        self.ocx.OnReceiveRealData[str, str, str].connect(
            self._OnReceiveRealData)
        self.ocx.OnReceiveChejanData[str, int, str].connect(
            self._OnReceiveChejanData)
        self.ocx.OnReceiveConditionVer[int, str].connect(
            self._OnReceiveConditionVer)
        self.ocx.OnReceiveTrCondition[str, str, str, int, int].connect(
            self._OnReceiveTrCondition)
        self.ocx.OnReceiveRealCondition[str, str, str, str].connect(
            self._OnReceiveRealCondition)

        self.timerSystem.setInterval(1000) 
        self.timerSystem.timeout.connect(self.onTimerSystemTimeout) 

        self.sigRealInfoArrived.connect(self.onRealInfoArrived)
  
    def isTradeAvailable(self):
        # 매수 가능 시간 체크 
        # 기본 정보를 얻기 위해서는 장 시작전 미리 동작을 시켜야 하고 매수를 위한 시간은 정확히 9시를 맞춤 (동시호가 시간의 매도 호가로 인해 매수 됨을 막기 위함)
        ret_vals= []
        current_time = self.currentTime.time()
        for start, stop in user_setting.AUTO_TRADING_OPERATION_TIME:
            start_time =  datetime.time(
                            hour = start[0],
                            minute = start[1] )
            stop_time =   datetime.time( 
                            hour = stop[0],
                            minute = stop[1] )
            if( current_time >= start_time and current_time <= stop_time ):
                ret_vals.append(True)
            else:
                ret_vals.append(False)
                pass

        # 하나라도 True 였으면 거래 가능시간임  
        if( ret_vals.count(True) ):
            return True
        else:
            return False
        pass

    @pyqtSlot()
    def mainStateEntered(self):
        pass

    @pyqtSlot()
    def stockCompleteStateEntered(self):
        print(util.whoami())
        self.sigStateStop.emit()
        pass

    @pyqtSlot()
    def initStateEntered(self):
        print(util.whoami())
        self.sigInitOk.emit()
        pass

    @pyqtSlot()
    def disconnectedStateEntered(self):
        # print(util.whoami())
        if( self.getConnectState() == 0 ):
            self.commConnect()
            QTimer.singleShot(90000, self.sigTryConnect)
            pass
        else:
            self.sigConnected.emit()
            
    @pyqtSlot()
    def connectedStateEntered(self):
        # get 계좌 정보

        account_cnt = self.getLoginInfo("ACCOUNT_CNT")
        acc_num = self.getLoginInfo("ACCNO")
        user_id = self.getLoginInfo("USER_ID")
        user_name = self.getLoginInfo("USER_NAME")
        keyboard_boan = self.getLoginInfo("KEY_BSECGB")
        firewall = self.getLoginInfo("FIREW_SECGB")
        print("account count: {}, acc_num: {}, user id: {}, " 
              "user_name: {}, keyboard_boan: {}, firewall: {}"
              .format(account_cnt, acc_num, user_id, user_name, 
                      keyboard_boan, firewall))

        self.account_list = (acc_num.split(';')[:-1])

        print(util.whoami() + 'account list ' + str(self.account_list))

        # 코스피 , 코스닥 종목 코드 리스트 얻기 
        result = self.getCodeListByMarket('0')
        self.kospiCodeList = tuple(result.split(';'))
        result = self.getCodeListByMarket('10')
        self.kosdaqCodeList = tuple(result.split(';'))
        pass

    @pyqtSlot()
    def systemStateEntered(self):
        pass

    @pyqtSlot()
    def initSystemStateEntered(self):
        # 체결정보 로드 
        if( os.path.isfile(CHEGYEOL_INFO_FILE_PATH) == True ):
            with open(CHEGYEOL_INFO_FILE_PATH, 'r', encoding='utf8') as f:
                file_contents = f.read()
                self.chegyeolInfo = json.loads(file_contents)

            # 한번 수익난 종목 매수 금지 
            current_date = self.currentTime.date().strftime("%y%m%d")
            if( current_date in self.chegyeolInfo):
                for line_str in self.chegyeolInfo[current_date]:
                    maedo_type = line_str.split("|")[3].strip()
                    jongmok_code = line_str.split("|")[4].strip()
                    if( '수동' in maedo_type ):
                        if( jongmok_code not in self.maesuProhibitCodeList ):
                            self.maesuProhibitCodeList.append(jongmok_code)

        if( os.path.isfile(JANGO_INFO_FILE_PATH) == True ):
            with open(JANGO_INFO_FILE_PATH, 'r', encoding='utf8') as f:
                file_contents = f.read()
                self.jangoInfoFromFile = json.loads(file_contents)

        # get 조건 검색 리스트
        self.getConditionLoad()
        self.timerSystem.start()
        pass

    @pyqtSlot()
    def requestingJangoSystemStateEntered(self):
        # print(util.whoami() )
        # 계좌 정보 조회 
        self.requestOpw00018(self.account_list[0], "0")
        pass 

    @pyqtSlot()
    def waitingTradeSystemStateEntered(self):
        # 장시작 전에 조건이 시작하도록 함 
        self.sigSelectCondition.emit()       

        # 반환값 : 조건인덱스1^조건명1;조건인덱스2^조건명2;…;
        # result = '조건인덱스1^조건명1;조건인덱스2^조건명2;'
        result = self.getConditionNameList()
        searchPattern = r'(?P<index>[^\/:*?"<>|;]+)\^(?P<name>[^\/:*?"<>|;]+);'
        fileSearchObj = re.compile(searchPattern, re.IGNORECASE)
        findList = fileSearchObj.findall(result)
        
        tempDict = dict(findList)
        print(tempDict)
        

        condition_name_screenNo_dict = {}
        for number, condition in tempDict.items():
            condition_name_screenNo_dict[condition] = [kw_util.sendConditionScreenNo + '{}'.format(int (number)), number]
        
        start_info_list = []
        start_name_list = []


        # 모든 리스트 종료 후 start 하도록 함
        for name, info in condition_name_screenNo_dict.items():

            if (name == self.current_condition_name ):
                start_info_list.append(info)
                start_name_list.append(name)
            else: 
                if( '이탈' in name ):
                    start_info_list.append(info)
                    start_name_list.append(name)
                else:
                    print("stop condition " + name + ", screen_no: " + info[0] + ", nIndex " + '{}'.format(int(info[1]) ) )
                    self.sendConditionStop( info[0], name, int(info[1]) )
                    self.disconnectRealData(info[0])
                    pass

        self.conditionOccurList.clear()
        self.conditionRemoveList.clear()

        for count in range(len(start_info_list)):
            print("start condition " + start_name_list[count] + ", screen_no: " + start_info_list[count][0] + ", nIndex " + '{}'.format(int(start_info_list[count][1])) )
            self.sendCondition( start_info_list[count][0], start_name_list[count], int(start_info_list[count][1]) , 1) 

        pass


    @pyqtSlot()
    def standbySystemStateEntered(self):
        print(util.whoami() )
        self.makeJangoInfoFile()
        # 연속으로 최대 5개 가능하므로 5개까지 기다림 
        QTimer.singleShot( TR_TIME_LIMIT_MS * 5, self.sigStartProcessBuy)
        pass

    @pyqtSlot()
    def terminatingSystemStateEntered(self):
        print(util.whoami() )
        pass


    @pyqtSlot()
    def processBuyStateEntered(self):
        pass
     
    @pyqtSlot()
    def initProcessBuyStateEntered(self):
        print(util.whoami())
        pass

    @pyqtSlot()
    def standbyProcessBuyStateEntered(self):
        # print(util.whoami() )
        # 바로 signal 발생하는 경우 너무 빨리 많이 돌아감 
        # 조건 발생 종목까지 봐야 하므로 종목이 많을 수 있고  많으면 100ms 너무 느림 조건  
        # 1초에 같은 종목 두번은 돌릴수 있어야 함 
        QTimer.singleShot(20, self.sigRequestEtcInfo)


    @pyqtSlot()
    def requestEtcInfoProcessBuyStateEntered(self):
        #장중에 한번만 얻어도 되는 정보를 요청할 때  
        # print(util.whoami() )

        # 조건 발생 리스트 검색 
        for jongmok_code in self.conditionRemoveList:
            self.removeConditionOccurList(jongmok_code)
        self.conditionRemoveList.clear()

        self.refreshRealRequest()

        jongmok_info_dict = self.getConditionOccurList()

        if( jongmok_info_dict == None ):
            self.sigError.emit()
            return 

        jongmok_code = jongmok_info_dict['종목코드']
        jongmok_name = jongmok_info_dict['종목명'] 

        ##########################################################################################################
        # key_day_candle = '일{}봉'.format(user_setting.MAX_SAVE_CANDLE_COUNT)
        # key_day_low_candle = '{}일봉중저가'.format(user_setting.STOP_LOSS_CALCULATE_DAY)


        ##########################################################################################################
        # 기본정보 요청 
        # 기준가의 경우 당일 상장 종목의 경우 공백일 수 있음 
        # if ( '기준가' not in jongmok_info_dict ):
        #     #기본 정보 여부 확인 
        #     self.requestOpt10001(jongmok_code)
        #     print("request {}".format(jongmok_name) )


        ##########################################################################################################
        # 일봉 정보 요청 
        # if(key_day_candle not in jongmok_info_dict):
        #     #일봉 정보 여부 확인 
        #     self.requestOpt10081(jongmok_code)
        #     print("request {}".format(jongmok_name) )


        ##########################################################################################################
        # 업봉 분봉 정보 요청 (첫접속시 한번만 )
        if( '분봉' not in self.upjongInfo['코스피'] ):
            if( self.requestOpt20005('001') == False ):
                self.sigError.emit()
            else:
                self.sigWaitTr.emit()
        elif( '분봉' not in self.upjongInfo['코스닥'] ): 
            if( self.requestOpt20005('101') == False ):
                self.sigError.emit()
            else:
                self.sigWaitTr.emit()
        else:
            self.sigDetermineBuy.emit()
        pass

    @pyqtSlot()
    def determineBuyProcessBuyStateEntered(self):
        jongmok_info_dict = self.getConditionOccurList()

        # 조건 검색에 걸린 종목도 같이 리스트업 됨
        if( jongmok_info_dict != None ):
            # 기본, 실시간 정보 등 필요  정보가 없는 경우 매수 금지 
            if( 
                jongmok_info_dict.get('등락율', '') == '' 
                or jongmok_info_dict.get('호가시간', '') == '' 
                or self.upjongInfo['코스피'].get('현재가', '') ==''
                or self.upjongInfo['코스닥'].get('현재가', '') ==''
                ):
                self.shuffleConditionOccurList()
                self.sigNoWaitTr.emit()
                # print("1", end= '')
                return
        else:
            self.sigNoWaitTr.emit()
            # print("3", end= '')
            return

        if( self.isTradeAvailable() == False ):
            # print("4", end= '')
            self.sigNoWaitTr.emit()
            return


        is_log_print_enable = False
        return_vals = []
        printLog = ''
            
        jongmok_code = jongmok_info_dict['종목코드']
        jongmok_name = jongmok_info_dict['종목명']

        jongmok_jang_type = ''
        if( jongmok_code in self.kosdaqCodeList ):
            jongmok_jang_type = 'kosdaq'
        else:
            jongmok_jang_type = 'kospi'

        # 매도 호가기준 
        current_price = abs(int(jongmok_info_dict['(최우선)매도호가']))
        maesuHoga1 = abs(int(jongmok_info_dict['(최우선)매수호가']))
        open_price = abs(int(jongmok_info_dict['시가']))
        high_price = abs(int(jongmok_info_dict['고가']))
        low_price = abs(int(jongmok_info_dict['저가']))
        total_maedohoga_amount = abs(int(jongmok_info_dict['매도호가총잔량']))
        total_maesuhoga_amount = abs(int(jongmok_info_dict['매수호가총잔량']))
        maesuHoga1_amount = int(jongmok_info_dict['매수호가수량1'])
        maesuHoga2_amount = int(jongmok_info_dict['매수호가수량2'])
        maedoHoga1_amount = int(jongmok_info_dict['매도호가수량1'])
        maedoHoga2_amount = int(jongmok_info_dict['매도호가수량2'])


        # 호가 정보는 문자열로 기준가 대비 + , - 값이 붙어 나옴 

        ##########################################################################################################
        # 최근 매수가/분할 매수 횟수  정보 생성
        bunhal_maesu_list =  []
        maesu_count = 0 

        if( jongmok_code in self.jangoInfo):
            bunhal_maesu_list = self.jangoInfo[jongmok_code].get('분할매수이력', [])
            maesu_count = len(bunhal_maesu_list)

        ##########################################################################################################
        # 전일 종가를 얻기 위한 기준가 정보 생성
        gijunga = int(self.GetMasterLastPrice(jongmok_code))

        ##########################################################################################################
        # 제외 종목인지 확인 
        if( jongmok_code in user_setting.EXCEPTION_LIST ):
            printLog += "(제외종목)"
            return_vals.append(False)

        ##########################################################################################################
        # 최대 보유 할 수 있는 종목 보유수를 넘었는지 확인 
        jango_jongmok_code_list = self.jangoInfo.keys()

        for exception_jongmok_code in user_setting.EXCEPTION_LIST :
            if( exception_jongmok_code in jango_jongmok_code_list ):
                jango_jongmok_code_list.remove(exception_jongmok_code)
        if( len(jango_jongmok_code_list) < user_setting.MAX_STOCK_POSSESION_COUNT ):
            pass
        else:
            if( jongmok_code not in self.jangoInfo):
                printLog += "(종목최대보유중)"
                return_vals.append(False)
        pass


        ##########################################################################################################
        # 데이트레이딩 중지 
        day_trading_end_time = datetime.time( hour = 14, minute = 59 )

        if( self.current_condition_name != '장후반' and  self.currentTime.time() > day_trading_end_time ):
            printLog += "(데이트레이딩종료)"
            return_vals.append(False)

        

        ##########################################################################################################
        # 업종 추세가 좋지 않은 경우 매수 금지 장이 좋지 않은 경우 매수 금지 

        _upjong_20_candle_avr = 0

        key_upjong_name = ''

        if( jongmok_jang_type == 'kospi'):
            key_upjong_name = '코스피'
        else:
            key_upjong_name = '코스닥'

        current_upjong_price = abs(round(float(self.upjongInfo[key_upjong_name]['현재가']), 2))
        candle_list = self.upjongInfo[key_upjong_name]['분봉']
        price_list = [ abs(round( float(item.split(':')[0]), 2) ) for item in candle_list ]
        # 분봉데이터는 소숫점 둘째자리까지 표현되는데 * 100 한 값의 문자열임 
        _upjong_20_candle_avr = round( ( (sum(price_list[1:20]) / 100)  + current_upjong_price )/ 20 , 2)

        # if( _upjong_20_candle_avr > current_upjong_price
        #     and self.current_condition_name !='장후반'):
        #     printLog += '(업종추세하락)'
        #     return_vals.append(False)


        ##########################################################################################################
        #  추가 매수 횟수 제한   
        if( maesu_count < user_setting.BUNHAL_MAESU_LIMIT ):
            pass
        else:
            printLog += '(분할매수한계)'
            return_vals.append(False)


        if( jongmok_code in self.jangoInfo ):
            if( '매도중' in self.jangoInfo[jongmok_code]):
                printLog += '(매도중)'
                return_vals.append(False)


        ##########################################################################################################
        # 불필요한 추가 매수를 막기 위함인 경우 
        if( self.maesuProhibitCodeList.count(jongmok_code) == 0 ):
            pass
        else:
            printLog += '(거래금지종목)'
            return_vals.append(False)

        ##########################################################################################################
        # 종목 등락율을 조건 적용 
        #  +, - 붙는 소수이므로 float 으로 먼저 처리 
        updown_percentage = float(jongmok_info_dict['등락율'] )
        if( updown_percentage > 0 and updown_percentage < 26 ):
            pass
        else:
            printLog += '(종목등락율미충족: 등락율 {0})'.format(updown_percentage)
            return_vals.append(False)
        pass

        ##########################################################################################################
        # 첫 매수시만 적용되는 조건 
        if( jongmok_code not in self.jangoInfo ):
            # 시간제약
            # 장 시작시 첫봉은 동시호가 적용이므로 제외, 그 후 1봉은 봐야 되므로 그 시간 이후 매수 
            start_time =   datetime.time( hour = 9, minute = user_setting.REQUEST_MINUTE_CANDLE_TYPE * 2) 
            stop_time =   datetime.time( hour = 9, minute = 30) 
            stop_end_time =   datetime.time( hour = 13, minute = 30) 
            if( self.currentTime.time() < start_time
                ):
                # print("{} {} ".format(util.cur_time(),  jongmok_name), end= '')
                printLog += '(매수시간미충족)'
                return_vals.append(False)
            pass

            # if( self.currentTime.time() > stop_time
            #     and self.currentTime.time() < stop_end_time
            #     ):
            #     # print("{} {} ".format(util.cur_time(),  jongmok_name), end= '')
            #     printLog += '(매수시간미충족)'
            #     return_vals.append(False)
            # pass

            # stoploss 용 실시간 조건 리스트 종목에 걸린 경우
            for jongmok_list in self.conditionStoplossList.values(): 
                if( jongmok_code in jongmok_list):
                    # print("{} {} ".format(util.cur_time(),  jongmok_name), end= '')
                    printLog += '(매수조건미충족)'
                    return_vals.append(False)
                    break




        ##########################################################################################################
        # 추가 매수시만 적용되는 조건 
        else:
            ##########################################################################################################
            # 최근 매입가 대비 비교하여 추매 
            time_span = datetime.timedelta(days = 1)
            _yesterday_date = (self.currentTime - time_span).date()
            _today_date = self.currentTime.date()

            current_jango  = self.jangoInfo[jongmok_code]

            bunhal_maedo_info_list = current_jango.get('분할매도이력', [])  
            bunhal_maesu_info_list = current_jango.get('분할매수이력', [])  
            maeipga = int(current_jango['매입가'])

            bunhal_maedo_count = len(bunhal_maedo_info_list)
            bunhal_maesu_count = len(bunhal_maesu_info_list)

            first_bunhal_maesu_time_str = bunhal_maesu_info_list[0].split(':')[0] #날짜:가격:수량 
            first_maeip_price = int(bunhal_maesu_info_list[0].split(':')[1]) #날짜:가격:수량 

            last_bunhal_maesu_time_str = bunhal_maesu_info_list[-1].split(':')[0]  #날짜:가격:수량 
            last_maeip_price = int(bunhal_maesu_info_list[-1].split(':')[1]) #날짜:가격:수량 

            first_bunhal_maesu_date = datetime.datetime.strptime( first_bunhal_maesu_time_str, '%Y%m%d%H%M%S').date()
            last_bunhal_maesu_date = datetime.datetime.strptime( last_bunhal_maesu_time_str, '%Y%m%d%H%M%S').date()


            if( _yesterday_date >= first_bunhal_maesu_date 
                and _today_date > last_bunhal_maesu_date):
                # 어제 이전부터 매수했고 금일 추가 매수 된적이 없는 경우 
                #스윙종목
                # if(  current_price > last_maeip_price * 1.015
                #     ):
                #     pass            
                # else:
                #     # printLog += '(추매조건미충족)'
                return_vals.append(False)

            else: 
                # 당일 추가 매수 종목 

                first_bunhal_stoploss_percent = 1.03

                if( current_price > last_maeip_price
                    # and bunhal_maesu_count == 1 
                    and current_price > last_maeip_price * first_bunhal_stoploss_percent
                    ):
                    pass
                else:
                    # printLog += '(추매조건미충족)'
                    return_vals.append(False)
                pass

            temp = '({} {})' .format( 
                    jongmok_name,  
                    current_price )
            # print( util.cur_time_msec() , temp)
            printLog += temp
            pass

        ##########################################################################################################
        # 매수  
        # 매도 호가가 0인경우 상한가임 
        if( return_vals.count(False) == 0 and current_price != 0  ):
            qty = 0
            if( user_setting.TEST_MODE == True ):
                qty = 1 
            else:
                # 기존 테스트 매수 수량을 조절하기 위함 
                if( jongmok_code in self.jangoInfo):

                    first_chegyeol_time_str = ""
                    if( len(bunhal_maesu_list) ):
                        first_chegyeol_time_str = bunhal_maesu_list[0].split(':')[0] # 날짜:가격:수량

                    if( first_chegyeol_time_str != ''):
                        base_time = datetime.datetime.strptime("20200806010101", "%Y%m%d%H%M%S") 

                        first_maesu_time = datetime.datetime.strptime(first_chegyeol_time_str, "%Y%m%d%H%M%S") 
                        total_price = user_setting.MAESU_TOTAL_PRICE[maesu_count] 
                        if( base_time < first_maesu_time ):
                            qty = int(total_price / current_price )  + 1 #  약간 오버하게 삼 
                            pass
                        else:
                            qty = int(100000 / current_price) + 1
                    else:
                        pass
                else:
                    # 신규 매수 
                    total_price = user_setting.MAESU_TOTAL_PRICE[maesu_count] 
                    qty = int(total_price / current_price )  + 1

                    if( self.current_condition_name == '장후반'):
                        qty = qty/3

            # result = self.sendOrder("buy_" + jongmok_code, kw_util.sendOrderScreenNo, 
            #                     objKiwoom.account_list[0], kw_util.dict_order["신규매수"], jongmok_code, 
            #                     qty, 0 , kw_util.dict_order["시장가"], "")

            current_price = kw_util.getHogaPrice(current_price, 1, jongmok_jang_type)
            result = self.sendOrder("buy_" + jongmok_code, kw_util.sendOrderScreenNo, 
                                objKiwoom.account_list[0], kw_util.dict_order["신규매수"], jongmok_code, 
                                qty, current_price , kw_util.dict_order["지정가"], "")

            self.maesuProhibitCodeList.append(jongmok_code)

            printLog = '{} **** [매수수량: {}, 매수가: {}, 매수호가 {}, 매도호가수량1: {}, 매도호가수량2: {}, 매수횟수: {}] ****'.format(
                jongmok_name,
                qty,
                current_price, 
                maesuHoga1, 
                maedoHoga1_amount, 
                maedoHoga2_amount,
                maesu_count
                )  
            print( printLog )

            util.save_log(printLog, '매수', folder = "log")
            self.sigWaitTr.emit()
        else:
            # print(printLog)
            self.sigNoWaitTr.emit()
            pass

        self.shuffleConditionOccurList()

        if( is_log_print_enable ):
            util.save_log(printLog, '조건진입', folder = "log")
        pass
     
    @pyqtSlot()
    def waitingTRlimitProcessBuyStateEntered(self):
        # print(util.whoami())
        # TR 은 개당 3.7 초 제한 
        # 5연속 조회시 17초 대기 해야함 
        # print(util.whoami() )
        QTimer.singleShot(TR_TIME_LIMIT_MS,  self.sigTrWaitComplete)
        pass 

    @pyqtSlot()
    def finalStateEntered(self):
        print(util.whoami())
        util.save_log('', subject= '', folder='log')
        util.save_log('', subject= '', folder='log')
        util.save_log('', subject= '', folder='log')
        util.save_log('', subject= '', folder='log')
        util.save_log('', subject= '', folder='log')
        util.save_log('', subject= '', folder='log')
        util.save_log('', subject= '', folder='log')
        util.save_log('', subject= '', folder='log')
        util.save_log('', subject= '', folder='log')

        # import subprocess
        # subprocess.call(["shutdown", "-s", "-t", "500"])
        pass


    def sendorder_multi(self, rQName, screenNo, accNo, orderType, code, qty, price, hogaGb, orgOrderNo):
        def inner():
            self.sendOrder(rQName, screenNo, accNo, orderType, code, qty, price, hogaGb, orgOrderNo)
        return inner


    def printStockInfo(self, jongmok_code = 'all'):
        if( jongmok_code == 'all'):
            print(json.dumps(self.jangoInfo, ensure_ascii= False, indent =2, sort_keys = True))
        else:
            print(json.dumps(self.jangoInfo[jongmok_code], ensure_ascii= False, indent =2, sort_keys = True))
        pass
    
    def printChegyeolInfo(self, current_date = 'all'):
        if( current_date == 'all' ):
            print(json.dumps(self.chegyeolInfo, ensure_ascii= False, indent = 2, sort_keys = True))
        elif( current_date == ''):
            current_date = self.currentTime.date().strftime("%y%m%d")
            if( current_date in self.chegyeolInfo):
                print(json.dumps(self.chegyeolInfo[current_date], ensure_ascii= False, indent = 2, sort_keys = True))

    def printYupjongInfo(self):
        print(json.dumps(self.upjongInfo, ensure_ascii= False, indent =2, sort_keys = True))

    # 주식 잔고정보 요청 
    def requestOpw00018(self, account_num, sPrevNext):
        self.setInputValue('계좌번호', account_num)
        self.setInputValue('비밀번호', '') #  사용안함(공백)
        self.setInputValue('비빌번호입력매체구분', '00')
        self.setInputValue('조회구분', '1')

        # 연속 데이터 조회해야 하는 경우 
        if( sPrevNext == "2" ):
            ret = self.commRqData(account_num, "opw00018", 2, kw_util.sendAccountInfoScreenNo) 
        else:
            ret = self.commRqData(account_num, "opw00018", 0, kw_util.sendAccountInfoScreenNo) 

        errorString = None
        if( ret != 0 ):
            errorString =   account_num + " commRqData() " + kw_util.parseErrorCode(str(ret))
            print(util.whoami() + errorString ) 
            util.save_log(errorString, util.whoami(), folder = "log" )
            return False
        return True

    # 주식 잔고 정보 #rQName 의 경우 계좌 번호로 넘겨줌
    def makeOpw00018Info(self, rQName):
        data_cnt = self.getRepeatCnt('opw00018', rQName)
        for cnt in range(data_cnt):
            info_dict = {}
            for item_name in kw_util.dict_jusik['TR:계좌평가잔고내역요청']:
                result = self.getCommData("opw00018", rQName, cnt, item_name)
                # 없는 컬럼은 pass 
                if( len(result) == 0 ):
                    continue
                if( item_name == '종목명'):
                    info_dict[item_name] = result.strip() 
                elif( item_name == '종목번호'):
                    info_dict[item_name] = result[1:-1].strip()
                elif( item_name == '수익률(%)'):
                    info_dict[item_name] = int(result) / 100
                else: 
                    info_dict[item_name] = int(result)
        
            jongmok_code = info_dict['종목번호']
            info_dict['업종'] = self.getMasterStockInfo(jongmok_code)
            
            if( jongmok_code not in self.jangoInfo.keys() ):
                self.jangoInfo[jongmok_code] = info_dict
            else:
                # 기존에 가지고 있는 종목이면 update
                self.jangoInfo[jongmok_code].update(info_dict)

        # print(self.jangoInfo)
        return True 

    # 주식 1일봉 요청 
    def requestOpt10081(self, jongmok_code):
        # print(util.cur_time_msec() )
        datetime_str = datetime.datetime.now().strftime('%Y%m%d')
        self.setInputValue("종목코드", jongmok_code)
        self.setInputValue("기준일자", datetime_str)    
        self.setInputValue('수정주가구분', '1')
        ret = self.commRqData(jongmok_code, "opt10081", 0, kw_util.sendGibonScreenNo) 
        errorString = None
        if( ret != 0 ):
            errorString = jongmok_code + " commRqData() " + kw_util.parseErrorCode(str(ret))
            print(util.whoami() + errorString ) 
            util.save_log(errorString, util.whoami(), folder = "log" )
            return False
        return True
        
    # 주식 일봉 데이터 생성  
    def makeOpt10081Info(self, rQName):
        # 조건 발생한 종목 상위 리스트의 정보를 얻기 위함 
        jongmok_info_dict = self.getConditionOccurList()
        if( jongmok_info_dict == None ):
            return False

        # 한번 읽으면 900개 읽힘 
        repeatCnt = self.getRepeatCnt("opt10081", rQName)
        jongmok_code = rQName
        # 거래되지 않는 종목의 경우 저가가 '' 값이 오기 때문에 1개는 채워둠 
        low_price_list = [999999999] 
        total_current_price_list = []

        # 일봉의 현재 봉도 포함 
        for i in range(min(repeatCnt, 200)): 
            for item_name in kw_util.dict_jusik['TR:일봉']:
                if( item_name == "저가"):
                    result = self.getCommData("opt10081", rQName, i, item_name)
                    if( i != 0 and result != '' and i <= user_setting.STOP_LOSS_CALCULATE_DAY):
                        # 첫번째는 당일이므로 제외 
                        low_price_list.append( abs(int(result)  ) )

                if( item_name == "현재가"):
                    result = self.getCommData("opt10081", rQName, i, item_name)
                    if( i != 0 and result != '' ):
                        # 첫번째는 당일이므로 제외 
                        total_current_price_list.append( abs(int(result)  ) )

        jongmok_info_dict['{}일봉중저가'.format(user_setting.STOP_LOSS_CALCULATE_DAY)] = min(low_price_list)
        jongmok_info_dict['일{}봉'.format(user_setting.MAX_SAVE_CANDLE_COUNT)] = total_current_price_list[0:user_setting.MAX_SAVE_CANDLE_COUNT]

        return True

    # 주식 분봉 tr 요청 
    def requestOpt10080(self, jongmok_code):
     # 분봉 tr 요청의 경우 너무 많은 데이터를 요청하므로 한개씩 수행 
        candle_type_str = "{}:{}분".format( user_setting.REQUEST_MINUTE_CANDLE_TYPE )

        self.setInputValue("종목코드", jongmok_code )
        self.setInputValue("틱범위", candle_type_str) 
        self.setInputValue("수정주가구분","1") 

        # rQName 을 데이터로 외부에서 사용
        ret = self.commRqData(jongmok_code , "opt10080", 0, kw_util.send5minScreenNo) 
        errorString = None
        if( ret != 0 ):
            errorString =  jongmok_code + " commRqData() " + kw_util.parseErrorCode(str(ret))
            print(util.whoami() + errorString ) 
            util.save_log(errorString, util.whoami(), folder = "log" )
            return False
        return True

    # 분봉 데이터 생성 
    def makeOpt10080Info(self, rQName):
        jongmok_code = rQName
        jongmok_info_dict = self.getConditionOccurList()
        if( jongmok_info_dict == None ):
            return False

        # 한번 읽으면 900개 읽힘 
        repeatCnt = self.getRepeatCnt("opt10080", rQName)

        total_current_price_list = []

        # 3분봉 기준 6.5 시간 * 20 = 130  
        for i in range(min(repeatCnt, user_setting.MAX_SAVE_CANDLE_COUNT)):
            line = []
            for item_name in kw_util.dict_jusik['TR:분봉']:
                result = self.getCommData("opt10080", rQName, i, item_name)
                if( item_name == "체결시간" ):
                    # 20191104145500 형식 
                    if( i == 0 ):
                        jongmok_info_dict['최근{}분봉체결시간'.format(user_setting.REQUEST_MINUTE_CANDLE_TYPE)] = result.strip()
                    line.append( result.strip() )
                else:
                    line.append( abs(int(result.strip()) ))
                    pass
            total_current_price_list.append( line )

        key_minute_candle = '{}분{}봉'.format(user_setting.REQUEST_MINUTE_CANDLE_TYPE, user_setting.MAX_SAVE_CANDLE_COUNT)
        jongmok_info_dict[key_minute_candle] = total_current_price_list

        if( jongmok_code in self.jangoInfo ):
            self.jangoInfo[jongmok_code][key_minute_candle] = jongmok_info_dict[key_minute_candle]

        
        return True

    # 업종 분봉 tr 요청 
    def requestOpt20005(self, yupjong_code):
        self.setInputValue("업종코드", yupjong_code )
        self.setInputValue("틱범위","3:5분") 
        self.setInputValue("수정주가구분","1") 
        ret = 0
        if( yupjong_code == '001'):
            ret = self.commRqData(yupjong_code , "opt20005", 0, kw_util.sendReqYupjongKospiScreenNo) 
        else:
            ret = self.commRqData(yupjong_code , "opt20005", 0, kw_util.sendReqYupjongKosdaqScreenNo) 
        
        errorString = None
        if( ret != 0 ):
            errorString =  yupjong_code + " commRqData() " + kw_util.parseErrorCode(str(ret))
            print(util.whoami() + errorString ) 
            util.save_log(errorString, util.whoami(), folder = "log" )
            return False
        return True

    # 업종 분봉 데이터 생성 
    def makeOpt20005Info(self, rQName):
        min_candle_list = []

        repeatCnt = self.getRepeatCnt("opt20005", rQName)

        for i in range(min(repeatCnt, user_setting.MAX_SAVE_CANDLE_COUNT)):
            line = []
            for item_name in kw_util.dict_jusik['TR:업종분봉']:
                result = self.getCommData("opt20005", rQName, i, item_name)
                line.append(result.strip())

                if( item_name == "체결시간" ):
                    line_item = '{}'.format( ':'.join(line))
                    min_candle_list.append(line_item)
        
        if( rQName == '001'):
            self.upjongInfo['코스피']['분봉'] = min_candle_list
        elif( rQName == '101'):
            self.upjongInfo['코스닥']['분봉'] = min_candle_list
        else:
            return False
        return True


    # 주식 기본 정보 요청  
    def requestOpt10001(self, jongmok_code):
        # print(util.cur_time_msec() )
        self.setInputValue("종목코드", jongmok_code)
        ret = self.commRqData(jongmok_code, "opt10001", 0, kw_util.sendGibonScreenNo) 
        errorString = None
        if( ret != 0 ):
            errorString = jongmok_code + " commRqData() " + kw_util.parseErrorCode(str(ret))
            print(util.whoami() + errorString ) 
            util.save_log(errorString, util.whoami(), folder = "log" )
            return False
        return True
        
    # 주식 기본 데이터 ( multi data 아님 )
    def makeOpt10001Info(self, rQName):

        jongmok_info_dict = self.getConditionOccurList()
        if( jongmok_info_dict == None):
            return False

        jongmok_code = rQName

        for item_name in kw_util.dict_jusik['TR:기본정보']:
            result = self.getCommData("opt10001", rQName, 0, item_name)
            if( jongmok_code in self.jangoInfo ):
                self.jangoInfo[jongmok_code][item_name] = result.strip()
            jongmok_info_dict[item_name] = result.strip()
        return True

    @pyqtSlot()
    def onTimerSystemTimeout(self):

        self.currentTime = datetime.datetime.now()

        jang_choban_start_time = datetime.time( hour = 8, minute = 0,  second = 30 )
        jang_choban_end_time = datetime.time( hour = 15, minute = 10 )
        jang_jungban_start_time = datetime.time( hour = 15, minute = 19 )

        current_time = self.currentTime.time()

        isConditionRefreshed = False

        if( current_time > jang_choban_start_time and current_time < jang_choban_end_time ):
            if( self.current_condition_name != '장초반' ):
                isConditionRefreshed = True
            self.current_condition_name = "장초반"
        elif( current_time > jang_jungban_start_time ):
            # if( self.current_condition_name != '장후반' ):
            #     isConditionRefreshed = True
            # self.current_condition_name = "장후반"
            pass
        else:
            # if( self.current_condition_name != '휴식' ):
            #     isConditionRefreshed = True
            # self.current_condition_name = "휴식"
            pass

        if(isConditionRefreshed == True):
            self.sigReselectCondition.emit()

        if( self.getConnectState() != 1 ):
            util.save_log("Disconnected!", "시스템", folder = "log")
            self.sigDisconnected.emit() 
        else:
            if( datetime.time(*TRADING_INFO_GETTING_TIME) <=  self.currentTime.time() ): 
                self.timerSystem.stop()
                self.make_excel(CHEGYEOL_INFO_EXCEL_FILE_PATH, self.chegyeolInfo)
                util.save_log("Stock Trade Terminate!\n\n\n\n\n", "시스템", folder = "log")
                pass
            else :
                pass
        pass

    @pyqtSlot()
    def quit(self):
        print(util.whoami())
        self.commTerminate()
        QApplication.quit()

    # 에러코드의 메시지를 출력한다.
    @pyqtSlot(int, result=str)
    def parseErrorCode(self, errCode):
        return kw_util.parseErrorCode(errCode)

    # event
    # 통신 연결 상태 변경시 이벤트
    # nErrCode가 0이면 로그인 성공, 음수면 실패
    def _OnEventConnect(self, errCode):
        print(util.whoami() + '{}'.format(errCode))
        if errCode == 0:
            self.sigConnected.emit()
        else:
            self.sigDisconnected.emit()

    # 수신 메시지 이벤트
    def _OnReceiveMsg(self, scrNo, rQName, trCode, msg):
        # print(util.whoami() + 'sScrNo: {}, sRQName: {}, sTrCode: {}, sMsg: {}'
        # .format(scrNo, rQName, trCode, msg))

        # [107066] 매수주문이 완료되었습니다.
        # [107048] 매도주문이 완료되었습니다
        # [571489] 장이 열리지않는 날입니다
        # [100000] 조회가 완료되었습니다
        printData =  'sScrNo: {}, sRQName: {}, sTrCode: {}, sMsg: {}'.format(scrNo, rQName, trCode, msg)

        # buy 하다가 오류 난경우 강제로 buy signal 생성  
        # buy 정상 메시지는 107066
        if( 'buy' in rQName and '107066' not in msg ):
            self.sigWaitTr.emit()

        # sell 하다가 오류 난경우 강제로 buy signal 생성  
        # sell 정상 메시지는 107066
        if( 'sell' in rQName and '107048' not in msg ):
            pass
            # self.sigWaitTr.emit()

        # sell 하다가 오류 난경우 
        # if( 'sell' in rQName and '매도가능수량' in msg ):
        #     self.sigWaitTr.emit()

        print(printData)
        util.save_log(printData, "시스템메시지", folder="log")
        pass

    # Tran 수신시 이벤트
    def _OnReceiveTrData(   self, scrNo, rQName, trCode, recordName,
                            prevNext, dataLength, errorCode, message,
                            splmMsg):
        # print(util.whoami() + 'sScrNo: {}, rQName: {}, trCode: {}, recordName: {}, prevNext {}, errorCode:{}, message:{}, splmMsg:{}' 
        # .format(scrNo, rQName, trCode, recordName, prevNext, errorCode, message, splmMsg ))

        if ( trCode == 'opw00018' ):
        # 게좌 정보 요청 rQName 은 계좌번호임 
            if( self.makeOpw00018Info(rQName) ):
                # 연속 데이터 존재 하는 경우 재 조회 
                if( prevNext  == "2" ) :
                    self.requestOpw00018(self.account_list[0], prevNext)
                else:
                    QTimer.singleShot(TR_TIME_LIMIT_MS,  self.sigRequestJangoComplete)

            else:
                self.sigError.emit()
            pass

        #주식 기본 정보 요청 rQName 은 개별 종목 코드임
        elif( trCode == "opt10001"):
            if( self.makeOpt10001Info(rQName) ):
                self.sigWaitTr.emit()
            else:
                self.sigError.emit()
            pass

        #주식 일봉 정보 요청 rqName 은 개별 종목 코드임  
        elif( trCode =='opt10081'):
            if( self.makeOpt10081Info(rQName) ):
                self.sigWaitTr.emit()
                pass
            else:
                self.sigError.emit()


        # 주식 분봉 정보 요청 rQName 개별 종목 코드  
        elif( trCode == "opt10080"):     
            if( self.makeOpt10080Info(rQName) ) :
                self.sigWaitTr.emit()
            else:
                self.sigError.emit()
            pass

        # 업종 분봉 rQName 업종 코드  
        elif( trCode == "opt20005"):     
            if( self.makeOpt20005Info(rQName) ) :
                self.sigWaitTr.emit()
                pass
            else:
                self.sigError.emit()
            pass

    # 실시간 시세 이벤트
    def _OnReceiveRealData(self, jongmok_code, realType, realData):
        # print(util.whoami() + 'jongmok_code: {}, {}, realType: {}'
        #         .format(jongmok_code, self.getMasterCodeName(jongmok_code),  realType))

        # 장전에도 주식 호가 잔량 값이 올수 있으므로 유의해야함 
        if( realType == "주식호가잔량"):
            # print(util.whoami() + 'jongmok_code: {}, realType: {}, realData: {}'
            #     .format(jongmok_code, realType, realData))

            self.makeRealDataInfo(jongmok_code, '실시간-{}'.format(realType) ) 

        #주식 체결로는 사고 팔기에는 반응이 너무 느림 
        elif( realType == "주식체결"):
            # print(util.whoami() + 'jongmok_code: {}, realType: {}, realData: {}'
            #     .format(jongmok_code, realType, realData))
            self.makeRealDataInfo(jongmok_code, '실시간-{}'.format(realType) ) 

            # WARNING: 장중에 급등으로 동시 호가진행되는 경우에 대비하여 체결가 정보 발생했을때만 stoploss 진행함. 
            self.processStopLoss(jongmok_code)
            pass
        
        elif( realType == "주식시세"):
            # 장종료 후에 나옴 
            print(util.whoami() + 'jongmok_code: {}, realType: {}, realData: {}'
                .format(jongmok_code, realType, realData))
            pass
        
        elif( realType == "업종지수" ):
            # print(util.whoami() + 'jongmok_code: {}, realType: {}, realData: {}'
            #     .format(jongmok_code, realType, realData))
            result = '' 
            key_name = ''
            if( jongmok_code == '001'):
                key_name = '코스피'
            elif( jongmok_code == '101'):
                key_name = '코스닥'
            upjong = self.upjongInfo[key_name]

            for col_name in kw_util.dict_jusik['실시간-{}'.format(realType)]:
                result = self.getCommRealData(jongmok_code, kw_util.name_fid[col_name] ) 
                upjong[col_name] = result.strip()

            if( '분봉' in upjong ):
                # 분봉 정보는 소수점이 없고 실시간 정보는 소수점 둘째자리 표시되는 문자열임
                current_price_str = str(round(float(upjong['현재가']) * 100, 2) )
                current_chegyeol_time_str = upjong['체결시간']
                # 장마감후 '체결시간' 이 장마감 문자열로 옴
                if( current_chegyeol_time_str != '장마감'):
                    current_chegyeol_time = datetime.datetime.strptime(current_chegyeol_time_str, "%H%M%S").time().replace(second=0)
                else:
                    current_chegyeol_time = self.currentTime.time()

                last_chegyeol_time_str = upjong['분봉'][0].split(':')[1]
                last_chegyeol_time = datetime.datetime.strptime(last_chegyeol_time_str, "%Y%m%d%H%M%S")

                time_span = datetime.timedelta(minutes= 3)

                if( current_chegyeol_time >= (last_chegyeol_time + time_span).time().replace(second=0) ):
                    upjong['분봉'].insert(0, '{}:{}'.format(current_price_str, '19990101{}'.format( current_chegyeol_time_str) ) )
                    upjong['분봉'] = upjong['분봉'][0:40]
                    # print(self.upjongInfo[key_name])

            pass 

        
        elif( realType == '장시작시간'):
            # TODO: 장시작 30분전부터 실시간 정보가 올라오는데 이를 토대로 가변적으로 장시작시간을 가늠할수 있도록 기능 추가 필요 
            # 장운영구분(0:장시작전, 2:장종료전, 3:장시작, 4,8:장종료, 9:장마감)
            # 동시호가 시간에 매수 주문 
            result = self.getCommRealData(realType, kw_util.name_fid['장운영구분'] ) 
            if( result == '2'):
                self.sigTerminating.emit()
            elif( result == '4' ): # 장종료 후 5분뒤에 프로그램 종료 하게 함  
                QTimer.singleShot(300000, self.sigStockComplete)

            # print(util.whoami() + 'jongmok_code: {}, realType: {}, realData: {}'
            #     .format(jongmok_code, realType, realData))
        
            print(util.whoami() + 'jongmok_code: {}, realType: {}, realData: {}'
                .format(jongmok_code, realType, realData))
            pass

    def calculateSuik(self, jongmok_code, current_price, amount):
        current_jango = self.jangoInfo[jongmok_code]
        maeip_price = abs(int(current_jango['매입가']))

        maeip_danga = maeip_price + maeip_price* 0.00015
        maedo_danga = current_price - current_price * 0.00015 - current_price * 0.0025

        suik_price = round( (maedo_danga - maeip_danga) * amount , 2)
        current_jango['수익'] = suik_price 
        current_jango['수익율'] = round( ( (maedo_danga - maeip_danga)  / maeip_danga ) * 100 , 2) 
        pass

    # 개별 종목 관련 실시간 체결/호가 등 정보 update 
    def makeRealDataInfo(self, jongmok_code, real_data_type):
        #주식 호가 잔량 정보 요청 
        result_list = [] 
        for col_name in kw_util.dict_jusik[real_data_type]:
            result_list.append(self.getCommRealData(jongmok_code, kw_util.name_fid[col_name] ).strip())

        self.sigRealInfoArrived.emit(jongmok_code, real_data_type, result_list)

    def isMinCandleExist(self, current_jango):
        key_minute_candle = '{}분{}봉'.format(user_setting.REQUEST_MINUTE_CANDLE_TYPE, user_setting.MAX_SAVE_CANDLE_COUNT)
        if( key_minute_candle in current_jango
            and len( current_jango[key_minute_candle] ) == user_setting.MAX_SAVE_CANDLE_COUNT ):  # 분봉 정보 얻었는지 확인 
            return True
        else:
            return False

    def isDayCandleExist(self, current_jango):
        key_day_candle = '일{}봉'.format(user_setting.MAX_SAVE_CANDLE_COUNT ) 
        if( key_day_candle in current_jango
            # and len( current_jango[key_day_candle] ) == user_setting.MAX_SAVE_CANDLE_COUNT 
            ):  # 분봉 정보 얻었는지 확인 
            return True
        else:
            return False


    def processStopLoss(self, jongmok_code):
        jongmok_name = self.getMasterCodeName(jongmok_code)
        if( self.isTradeAvailable() == False ):
            print('-1', end = '')
            return
        
        # 예외 처리 리스트이면 종료 
        if( jongmok_code in user_setting.EXCEPTION_LIST ):
            # print('-2', end = '')
            return

        # 잔고에 없는 종목이면 종료 
        if( jongmok_code not in self.jangoInfo ):
            # print('-3', end = '')
            return 
        current_jango = self.jangoInfo[jongmok_code]

        bunhal_maedo_info_list = current_jango.get('분할매도이력', [])  
        bunhal_maesu_info_list = current_jango.get('분할매수이력', [])  

        bunhal_maedo_count = len(bunhal_maedo_info_list)
        bunhal_maesu_count = len(bunhal_maesu_info_list)

        first_bunhal_maesu_time_str = bunhal_maesu_info_list[0].split(':')[0] #날짜:가격:수량 
        first_maeip_price = int(bunhal_maesu_info_list[0].split(':')[1]) #날짜:가격:수량 

        last_bunhal_maesu_time_str = bunhal_maesu_info_list[-1].split(':')[0]  #날짜:가격:수량 
        last_maeip_price = int(bunhal_maesu_info_list[-1].split(':')[1]) #날짜:가격:수량 

        if( 
            '손절가' not in current_jango 
            or '매매가능수량' not in current_jango 
            or '매도호가총잔량' not in current_jango 
            or '현재가' not in self.upjongInfo['코스닥']
            or '현재가' not in self.upjongInfo['코스피']
            or '분봉' not in self.upjongInfo['코스닥']
            or '분봉' not in self.upjongInfo['코스피']
            ):
            print('-4', end = '')
            return

        # 중요: 매도 시 매도 호가가 빠지는 등의 이유로 가격의 왜곡이 생기므로 최우선 매수 호가를 기준 삼지 않는다. 
        # 손절시는 current_price 참고  
        # 익절시는 maesuHoga1 참고  (치고 올라가기 때문에  매수/매도 호가 괴리가 생기므로 매수호가 기준 )
        current_price = abs(int ( current_jango['현재가']))
        maesuHoga1 = abs(int ( current_jango['(최우선)매수호가']))

        # # 매수호가 기준 
        # if( jongmok_code in self.kospiCodeList ):
        #     current_price = kw_util.getHogaPrice(current_price, -1, 'kospi')
        # else: 
        #     current_price = kw_util.getHogaPrice(current_price, -1, 'kosdaq')

        jangosuryang = int( current_jango['매매가능수량'] )
        stop_plus = int(current_jango['이익실현가'])
        stop_loss = int(current_jango['손절가'])
        maeipga = int(current_jango['매입가'])

        updown_percentage = float(current_jango['등락율']) 

        time_span = datetime.timedelta(days = 1)

        total_maedohoga_amount = int(current_jango['매도호가총잔량'])
        total_maesuhoga_amount = int(current_jango['매수호가총잔량'])

        # 업종 정보 생성 
        jongmok_jang_type = ''
        if( jongmok_code in self.kosdaqCodeList ):
            jongmok_jang_type = 'kosdaq'
        else:
            jongmok_jang_type = 'kospi'

        _upjong_20_candle_avr = 0
        key_upjong_name = ''
        if( jongmok_jang_type == 'kospi'):
            key_upjong_name = '코스피'
        else:
            key_upjong_name = '코스닥'

        current_upjong_price = abs(round(float(self.upjongInfo[key_upjong_name]['현재가']), 2))
        candle_list = self.upjongInfo[key_upjong_name]['분봉']
        price_list = [ abs(round( float(item.split(':')[0]), 2) ) for item in candle_list ]
        # 분봉데이터는 소숫점 둘째자리까지 표현되는데 * 100 한 값의 문자열임 
        _upjong_20_candle_avr = round( ( (sum(price_list[1:20]) / 100)  + current_upjong_price )/ 20 , 2)

        ##########################################################################################################
        current_time_str = util.cur_time()
        server_hoga_time_str = current_jango['호가시간']
        expected_one_hoga_amount = round(total_maedohoga_amount)

        _yesterday_date = (self.currentTime - time_span).date()
        _today_date = (self.currentTime).date()

        _today_open_price = abs(int(current_jango['시가']))
        _today_close_price = abs(int(current_jango['현재가']))
        _today_low_price = abs(int(current_jango['저가']))
        _today_high_price = abs(int(current_jango['고가']) )
        _today_amount = abs(int(current_jango['누적거래량']))

        maesuHoga1_amount = abs(int(current_jango['매수호가수량1']))
        maesuHoga2_amount = abs(int(current_jango['매수호가수량2']))
        maedoHoga1_amount = abs(int(current_jango['매도호가수량1']))
        maedoHoga2_amount = abs(int(current_jango['매도호가수량2']))

        _yesterday_close_price = int(self.GetMasterLastPrice(jongmok_code))
        _percent = abs(float(current_jango['전일거래량대비(비율)'])) 
        _yesterday_amount = 0
        if(_percent != 0  ):
            _yesterday_amount = int( _today_amount / (_percent / 100) )

        maedo_type = ""

        first_bunhal_maesu_date_time = datetime.datetime.strptime( first_bunhal_maesu_time_str, '%Y%m%d%H%M%S').date()
        first_bunhal_stoploss_percent = 1.023

        if( _yesterday_date >= first_bunhal_maesu_date_time):
            ##########################################################################################################
            # 분할 매수 스윙 종목  
            stop_plus = 99999999
            maedo_type = "(스윙매수기본손절)"


        else:
            ##########################################################################################################
            # 당일 매수 종목 
            last_bunhal_maesu_date_time = datetime.datetime.strptime(last_bunhal_maesu_time_str, "%Y%m%d%H%M%S") 
            stop_plus = 9999999 
            bunhal_maedo_base_amount = 0
            maedo_type = "(당일매수기본손절)"


            # 본전 손절 적용 시간내로 안나오면 매도 
            time_span = datetime.timedelta( seconds= 10 )


        ########################################################################################

        isSell = False
        printData = jongmok_code + ' {0:20} '.format(jongmok_name) 

        # 손절 / 익절 계산 
        # 정리나, 손절의 경우 시장가로 팔고 익절의 경우 보통가로 팜 
        isSijanga = False
        sell_amount = 0

        # 20180410150510 팜스웰바이오 실시간 매수 호가가 0으로 나오는 경우 있음 
        if( stop_loss >= current_price and current_price > 0 ) :
            isSijanga = True
            isSell = True
        elif( stop_plus < maesuHoga1 ) :
            isSell = True 
        printData += maedo_type 
        

        printData +=    ' 손절가: {}, 이익실현가: {}, 매입가: {}, 잔고수량: {}, 호가시간: {}, 현재가: {}, 매수호가수량1: {}, 매수호가수량2: {}'.format(
                            stop_loss, stop_plus, maeipga , jangosuryang, 
                            server_hoga_time_str,
                            current_price, 
                            maesuHoga1_amount, maesuHoga2_amount 
                        )


        order_num = current_jango.get('주문번호', '')

        if( isSell == True ):
            # 키움 open api 에서는 시장가 정정 주문이 동작하지 않으므로 최대한 아래 값으로 넣는다. 
            low_price = current_price * 0.9
            low_price = kw_util.getHogaPrice(low_price, 0, jongmok_jang_type)
            result = self.sendOrder("sell_"  + jongmok_code, kw_util.sendOrderScreenNo, objKiwoom.account_list[0], kw_util.dict_order["매도정정"], 
                                jongmok_code, sell_amount, low_price , kw_util.dict_order["지정가"], order_num)

            util.save_log(printData, '매도', folder='log')
            print("S {} 잔고수량: {}, 매도타입: {}, 주문번호:{},  {}".format(
                jongmok_name, sell_amount, maedo_type, order_num, result),  sep= "")
            pass


        pass

    # 체결데이터를 받은 시점을 알려준다.
    # sGubun – 0:주문체결통보, 1:잔고통보, 3:특이신호
    # sFidList – 데이터 구분은 ‘;’ 이다.
    '''
    매수시 
    14:03:36.218242 _OnReceiveChejanData gubun: 0, itemCnt: 35, fidList: 9201;9203;9205;9001;912;913;302;900;901;902;903;904;905;906;907;908;909;910;911;10;27;28;914;915;938;939;919;920;921;922;923;949;10010;969;819
    * 14:03:36.250244 _OnReceiveChejanData gubun: 0, itemCnt: 35, fidList: 9201;9203;9205;9001;912;913;302;900;901;902;903;904;905;906;907;908;909;910;911;10;27;28;914;915;938;939;919;920;921;922;923;949;10010;969;819
    * 14:03:36.353244 _OnReceiveChejanData gubun: 1, itemCnt: 34, fidList: 9201;9001;917;916;302;10;930;931;932;933;945;946;950;951;27;28;307;8019;957;958;918;990;991;992;993;959;924;10010;25;11;12;306;305;970
    '''
    '''
    _OnReceiveChejanData gubun: 1, itemCnt: 27, fidList: 9201;9001;917;916;302;10;930;931;932;933;945;946;950;951;27;28;307;8019;957;958;918;990;991;992;993;959;924
    {'종목코드': 'A010050', '당일실현손익률(유가)': '0.00', '대출일': '00000000', '당일실현손익률(신용)': '0.00', '(최우선)매수호가': '+805', '당일순매수수량': '5', '총매입가': '4043', 
    '당일총매도손일': '0', '만기일': '00000000', '신용금액': '0', '당일실현손익(신용)': '0', '현재가': '+806', '기준가': '802', '계좌번호': ', '보유수량': '5', 
    '예수금': '0', '주문가능수량': '5', '종목명': '우리종금                                ', '손익율': '0.00', '당일실현손익(유가)': '0', '담보대출수량': '0', '924': '0', 
    '매입단가': '809', '신용구분': '00', '매도/매수구분': '2', '(최우선)매도호가': '+806', '신용이자': '0'}
    ''' 
    def _OnReceiveChejanData(self, gubun, itemCnt, fidList):
        # print(util.whoami() + 'gubun: {}, itemCnt: {}, fidList: {}'
        #         .format(gubun, itemCnt, fidList))
        if( gubun == "1"): # 잔고 정보
            # 잔고 정보에서는 매도/매수 구분이 되지 않음 

            jongmok_code = self.getChejanData(kw_util.name_fid['종목코드'])[1:]
            boyou_suryang = int(self.getChejanData(kw_util.name_fid['보유수량']))
            old_boyou_suryang = 0
            if( jongmok_code in self.jangoInfo ):
                old_boyou_suryang = self.jangoInfo[jongmok_code].get('보유수량', 0)
            jumun_ganeung_suryang = int(self.getChejanData(kw_util.name_fid['주문가능수량']))
            maeip_danga = int(self.getChejanData(kw_util.name_fid['매입단가']))
            jongmok_name= self.getChejanData(kw_util.name_fid['종목명']).strip()
            current_price = abs(int(self.getChejanData(kw_util.name_fid['현재가'])))
            current_amount = abs(int(self.getChejanData(kw_util.name_fid['당일순매수수량'])))
            maesuHoga1 = abs(int(self.getChejanData(kw_util.name_fid['(최우선)매수호가'])))
            maedoHoga1 = abs(int(self.getChejanData(kw_util.name_fid['(최우선)매도호가'])))


            #미체결 수량이 있는 경우 잔고 정보 저장하지 않도록 함 
            if( jongmok_code in self.michegyeolInfo):
                if( self.michegyeolInfo[jongmok_code]['미체결수량'] ):
                    return 
                else:
                    # 미체결 수량이 없으므로 정보 삭제 
                    del ( self.michegyeolInfo[jongmok_code] )

            # 아래 잔고 정보의 경우 TR:계좌평가잔고내역요청 필드와 일치하게 만들어야 함 
            current_jango = {}
            current_jango['보유수량'] = boyou_suryang
            current_jango['매매가능수량'] =  jumun_ganeung_suryang # TR 잔고에서 매매가능 수량 이란 이름으로 사용되므로 
            current_jango['매입가'] = maeip_danga
            current_jango['종목번호'] = jongmok_code
            current_jango['종목명'] = jongmok_name.strip()
            current_jango['업종'] = self.getMasterStockInfo(jongmok_code)

            # 매수  
            if( boyou_suryang > old_boyou_suryang ):
                if( boyou_suryang > old_boyou_suryang ):
                    chegyeol_info = util.cur_date_time('%Y%m%d%H%M%S') + ":" + str(maeip_danga) + ":" + str(current_amount)
                    if( jongmok_code not in self.jangoInfo):
                        # 첫매수
                        current_jango['분할매수이력'] = [chegyeol_info] 
                        current_jango['매수시최우선매수호가'] = maesuHoga1
                        self.jangoInfo[jongmok_code] = current_jango 
                    else:
                        # 분할매수
                        last_chegyeol_info = self.jangoInfo[jongmok_code]['분할매수이력'][-1]
                        last_price = int(last_chegyeol_info.split(':')[1])
                        if( last_price != current_price ):
                            chegyeol_info_list = self.jangoInfo[jongmok_code].get('분할매수이력', [])
                            chegyeol_info_list.append( chegyeol_info )
                            current_jango['분할매수이력'] = chegyeol_info_list
                        pass
                if( jongmok_code in self.maesuProhibitCodeList):
                    self.maesuProhibitCodeList.remove(jongmok_code)
            # 매도
            elif( boyou_suryang < old_boyou_suryang ):
                if( boyou_suryang == 0 ):
                    # 보유 수량이 0 인 경우 완전 매도 수행한 것임  
                    self.jangoInfo.pop(jongmok_code)
                    if( jongmok_code in self.jangoInfoFromFile):
                        self.jangoInfoFromFile.pop(jongmok_code)
                    self.removeConditionOccurList(jongmok_code)
                else:
                    # 분할매도
                    current_amount = old_boyou_suryang - boyou_suryang
                    chegyeol_info = util.cur_date_time('%Y%m%d%H%M%S') + ":" + str(current_price) + ":" + str(current_amount)

                    chegyeol_info_list = self.jangoInfo[jongmok_code].get('분할매도이력', [])  
                    chegyeol_info_list.append( chegyeol_info )
                    current_jango['분할매도이력'] = chegyeol_info_list
                    pass

                    if( '매도중' in self.jangoInfo[jongmok_code] ):
                        del self.jangoInfo[jongmok_code]['매도중']
            else:
                print('매도시 첫 잔고')
            # 매도로 다 팔아 버린 경우가 아니라면 
            if( jongmok_code in self.jangoInfo ):
                self.jangoInfo[jongmok_code].update(current_jango)
            self.makeJangoInfoFile()
            pass

        elif ( gubun == "0"):
            jumun_sangtae =  self.getChejanData(kw_util.name_fid['주문상태'])
            jongmok_code = self.getChejanData(kw_util.name_fid['종목코드'])[1:]
            michegyeol_suryang = int(self.getChejanData(kw_util.name_fid['미체결수량']))

            # 주문 상태 
            # 매수 시 접수(gubun-0) - 체결(gubun-0) - 잔고(gubun-1)     
            # 매도 시 접수(gubun-0) - 잔고(gubun-1) - 체결(gubun-0) - 잔고(gubun-1)   순임 
            # 미체결 수량 정보를 입력하여 잔고 정보 처리시 미체결 수량 있는 경우에 대한 처리를 하도록 함 
            if( jongmok_code not in self.michegyeolInfo):
                self.michegyeolInfo[jongmok_code] = {}
            self.michegyeolInfo[jongmok_code]['미체결수량'] = michegyeol_suryang

            if( jumun_sangtae == "체결"):
                # 매수 체결과 매도 체결 구분해야함 
                self.makeChegyeolInfo(jongmok_code, fidList)
                self.makeChegyeolInfoFile()
                pass
            elif ( jumun_sangtae == '접수'):
                jumun_number = self.getChejanData(kw_util.name_fid['주문번호'])
                # 매도 접수인 경우 
                if( jongmok_code in self.jangoInfo ):
                    print("sell: {} ordernumber: {} 접수 ".format( self.getMasterCodeName(jongmok_code), jumun_number ) )
                    self.jangoInfo[jongmok_code]['주문번호'] = jumun_number



            
            pass


    def makeChegyeolInfoFile(self):
        # print(util.whoami())
        with open(CHEGYEOL_INFO_FILE_PATH, 'w', encoding = 'utf8' ) as f:
            f.write(json.dumps(self.chegyeolInfo, ensure_ascii= False, indent= 2, sort_keys = True ))
        pass

    # 서버에서 얻을 수 없는 추가 정보 저장
    # 첫 잔고 정보 요청시 호출됨 
    # 매수, 매도후 체결 정보로 잔고 정보 올때 호출됨 
    def makeEtcJangoInfo(self, jongmok_code): 
        if( jongmok_code not in self.jangoInfo):
            return

        current_jango = self.jangoInfo[jongmok_code]

        # 분할매수없는 경우 저장된 데이터 파일로부터 읽어오고 파일로부터도 없으면 default 값 입력 
        if( '분할매수이력' not in current_jango ):
            current_jango['분할매수이력'] = ['29991212091234:2:1']
            if( jongmok_code in self.jangoInfoFromFile):
                current_jango['분할매수이력'] = self.jangoInfoFromFile[jongmok_code].get('분할매수이력', [])

        # 분할매도없는 경우 저장된 데이터 파일로부터 읽어옴
        if( '분할매도이력' not in current_jango ):
            current_jango['분할매도이력'] = []
            if( jongmok_code in self.jangoInfoFromFile):
                current_jango['분할매도이력'] = self.jangoInfoFromFile[jongmok_code].get('분할매도이력', [])

        # 주문번호없는 경우 저장된 데티어 파일로부터 읽어옴 
        if( '주문번호' not in current_jango ):
            current_jango['주문번호'] = []
            if( jongmok_code in self.jangoInfoFromFile):
                current_jango['주문번호'] = self.jangoInfoFromFile[jongmok_code].get('주문번호', [])

        # 총 매수 갯수 게산 
        bunhal_maesu_list = current_jango['분할매수이력']
        maesu_count = len(bunhal_maesu_list)

        # 손절/익절가 퍼센티지 계산 
        stop_loss_percent = user_setting.STOP_LOSS_PER_MAESU_COUNT[maesu_count -1]
        stop_plus_percent = user_setting.STOP_PLUS_PER_MAESU_COUNT[maesu_count -1]

        maeip_price = current_jango['매입가']

        # 기본 손절가 측정 
        gibon_stoploss = round( maeip_price *  (1 + (stop_loss_percent/ 100) ) , 2 )

        print("종목이름:{}, 기본손절:{}".format(self.getMasterCodeName(jongmok_code), gibon_stoploss))
 
        ###############################################################################################
        current_jango['손절가'] =  gibon_stoploss
        current_jango['이익실현가'] = round( maeip_price * (1 + (stop_plus_percent/100) ) , 2)

        self.jangoInfo[jongmok_code].update(current_jango)
        pass

    @pyqtSlot()
    def makeJangoInfoFile(self):
        print(util.whoami())

        # 기타 정보 업데이트
        for jongmok_code in self.jangoInfo.keys():
            self.makeEtcJangoInfo(jongmok_code)

        temp = copy.deepcopy(self.jangoInfo)
        # 불필요 필드 제거 
        for jongmok_code, contents in temp.items():
            for key in self.jango_remove_keys:
                if( key in contents):
                    del contents[key]

        with open(JANGO_INFO_FILE_PATH, 'w', encoding = 'utf8' ) as f:
            f.write(json.dumps(temp, ensure_ascii= False, indent= 2, sort_keys = True ))
        pass

        # self.makeInterestedStocksFile()

    @pyqtSlot()
    def makeInterestedStocksFile(self):
        print(util.whoami())

        temp = copy.deepcopy(self.conditionOccurList)
        # 불필요 필드 제거 
        # condition list 
        for item in temp:
            for key in self.jango_remove_keys:
                if( key in item):
                    del item[key]

        with open(INTERESTED_STOCKS_FILE_PATH, 'w', encoding = 'utf8' ) as f:
            f.write(json.dumps(temp, ensure_ascii= False, indent= 2, sort_keys = True ))
        pass

    def makeChegyeolInfo(self, jongmok_code, fidList):
        fids = fidList.split(";")
        printData = "" 
        info = [] 

        # 미체결 수량이 0 이 아닌 경우 다시 체결 정보가 올라 오므로 0인경우 처리 안함 
        michegyeol_suryung = int(self.getChejanData(kw_util.name_fid['미체결수량']).strip())
        if( michegyeol_suryung != 0 ):
            return
        nFid = kw_util.name_fid['매도매수구분']
        result = self.getChejanData(nFid).strip()
        maedo_maesu_gubun = '매도' if result == '1' else '매수'
        # 첫 매수시는 잔고 정보가 없을 수 있으므로 
        current_jango = self.jangoInfo.get(jongmok_code, {})
        bunhal_maesu_list = current_jango.get('분할매수이력', [])
        bunhal_maedo_list = current_jango.get('분할매도이력', [])


        #################################################################################################
        # 사용자 정의 컬럼 수익과 수익율 필드 채움 
        if( maedo_maesu_gubun == '매도'): 
            # 체결가를 통해 수익율 필드 업데이트 
            current_price = int(self.getChejanData(kw_util.name_fid['체결가']).strip())
            current_maedo_amount = int(self.getChejanData(kw_util.name_fid['주문수량']).strip())

            self.calculateSuik(jongmok_code, current_price, current_maedo_amount)

            # 매도시 체결정보는 수익율 필드가 존재 
            profit = current_jango.get('수익', '0')
            profit_percent = current_jango.get('수익율', '0' )
            maesu_count = len(bunhal_maesu_list)
            maedo_type = current_jango.get('매도중', '')
            if( maedo_type == ''):
                # 한번 수익난 종목 당일 매수 금지 
                self.maesuProhibitCodeList.append(jongmok_code)
                maedo_type = '(수동직접매도수행)'
            info.append('{0:>10}'.format(profit_percent))
            info.append('{0:>10}'.format(profit))
            info.append(' 매수횟수: {0:>1} '.format(maesu_count))
            info.append(' {0} '.format(maedo_type))


        elif( maedo_maesu_gubun == '매수') :  
            # 매수시 체결정보는 수익율 / 수익 필드가  
            info.append('{0:>10}'.format('0'))
            info.append('{0:>10}'.format('0'))
            # 체결시는 매수 횟수 정보가 업데이트 되지 않았기 때문에 +1 해줌  
            # 첫매수에 대한 처리도 함
            maesu_count = len(bunhal_maesu_list)
            info.append(' 매수횟수: {0:>1} '.format(maesu_count + 1))
            info.append(' {0} '.format('(매수매수매수매수)'))

        pass

        #################################################################################################
        # kiwoom api 체결 정보 필드 
        for col_name in kw_util.dict_jusik["체결정보"]:
            nFid = None
            result = ""
            if( col_name not in kw_util.name_fid ):
                continue

            nFid = kw_util.name_fid[col_name]

            if( str(nFid) in fids):
                result = self.getChejanData(nFid).strip()
                if( col_name == '종목코드'):
                    result = result[1:] 
                if( col_name == '체결가' ):
                    result = '{0:>10}'.format(result)
                if( col_name == '주문구분'):
                    result = '{0:<10}'.format(result)
                
                if( col_name == '체결량' or col_name == '미체결수량'):
                    result = '{0:>7}'.format(result)

                info.append(' {} '.format(result))
                printData += col_name + ": " + result + ", " 
    
        current_date = self.currentTime.date().strftime('%y%m%d')

        if( current_date not in self.chegyeolInfo) :
            self.chegyeolInfo[current_date] = [] 

        #################################################################################################
        # 매도시 매수 이력 정보 필드 
        if( maedo_maesu_gubun == '매도'): 
            info.append(' | '.join(current_jango['분할매수이력']))
            pass

        self.chegyeolInfo[current_date].append('|'.join(info))
        util.save_log(printData, "*체결정보", folder= "log")
        pass


    # 로컬에 사용자조건식 저장 성공여부 응답 이벤트
    # 0:(실패) 1:(성공)
    def _OnReceiveConditionVer(self, ret, msg):
        print(util.whoami() + 'ret: {}, msg: {}'
            .format(ret, msg))
        if ret == 1:
            self.sigGetConditionCplt.emit()

    # 조건검색 조회응답으로 종목리스트를 구분자(“;”)로 붙어서 받는 시점.
    # LPCTSTR sScrNo : 종목코드
    # LPCTSTR strCodeList : 종목리스트(“;”로 구분)
    # LPCTSTR strConditionName : 조건명
    # int nIndex : 조건명 인덱스
    # int nNext : 연속조회(2:연속조회, 0:연속조회없음)
    def _OnReceiveTrCondition(self, scrNo, codeList, conditionName, index, next):
        # print(util.whoami() + 'scrNo: {}, codeList: {}, conditionName: {} '
        # 'index: {}, next: {}'
        # .format(scrNo, codeList, conditionName, index, next ))
        codes = codeList.split(';')[:-1]
        # 마지막 split 결과 None 이므로 삭제 
        if( '이탈' not in conditionName ):
            for code in codes:
                print('condition occur list add code: {} '.format(code) + self.getMasterCodeName(code))
                self.addConditionOccurList(code)

    # 편입, 이탈 종목이 실시간으로 들어옵니다.
    # strCode : 종목코드
    # strType : 편입(“I”), 이탈(“D”)
    # streonditionName : 조건명
    # strConditionIndex : 조건명 인덱스
    def _OnReceiveRealCondition(self, code, type, conditionName, conditionIndex):

        if ( '이탈' in conditionName ):
            # print('{} {}, code: {}, 종목이름: {},  type: {},  conditionIndex: {}'
            #     .format(util.cur_time_msec(), conditionName, code, self.getMasterCodeName(code), type, conditionIndex ))
            key_name = '3분'
            if( '이탈1'  == conditionName ):
                key_name = '1분'
                pass
            elif( '이탈3' == conditionName ):
                key_name = '3분'
                pass
            elif( '이탈15' == conditionName ):
                key_name = '15분'
                pass
            elif( '이탈30' == conditionName ):
                key_name = '30분'
                pass

            if ( type == 'I' ):
                if( code not in self.conditionStoplossList[key_name]):
                    self.conditionStoplossList[key_name].append(code)
            else:
                if( code in self.conditionStoplossList[key_name]):
                    self.conditionStoplossList[key_name].remove(code)
                pass

        else:
            if ( type == 'I' ):
                print('+{} {}'
                    .format(util.cur_time_msec(), self.getMasterCodeName(code) ))
                self.addConditionOccurList(code) # 조건 발생한 경우 해당 내용 list 에 추가  
            else:
                print('-{} {}'
                    .format(util.cur_time_msec(), self.getMasterCodeName(code) ))
                self.conditionRemoveList.append(code)
                pass


    def addConditionOccurList(self, jongmok_code):
        #발생시간, 종목코드,  종목명
        jongmok_name = self.getMasterCodeName(jongmok_code)
        ret_vals = []

        # 중복 제거 
        for item_dict in self.conditionOccurList:
            if( jongmok_code == item_dict['종목코드'] ):
                ret_vals.append(True)
            
        if( ret_vals.count(True) ):
            pass
        else:
            self.conditionOccurList.append( {'종목명': jongmok_name, '종목코드': jongmok_code} )
            self.sigConditionOccur.emit()
        pass
    
    def removeConditionOccurList(self, jongmok_code):
        for item_dict in self.conditionOccurList:
            if( item_dict['종목코드'] == jongmok_code ):
                self.conditionOccurList.remove(item_dict)
                break
        pass

    def getConditionOccurList(self):
        if( len(self.conditionOccurList ) ):
            return self.conditionOccurList[0]
        else:
            return None
        pass
    
    def getCodeListConditionOccurList(self):
        items = []
        for item_dict in self.conditionOccurList:
            items.append(item_dict['종목코드'])
        return items

    # 실시간 체결처리하기 위함
    def setRealData(self, real_data_type, item_dict, result_list):
        jongmok_code = ""

        if( '주식호가잔량' in real_data_type):
            if( '종목번호' in item_dict):
                jongmok_code = item_dict['종목번호']
            if( '종목코드' in item_dict):
                jongmok_code = item_dict['종목코드']

            # 실시간 데이터 대입 
            for index, col_name in enumerate(kw_util.dict_jusik[real_data_type]) :
                item_dict[col_name] = result_list[index]
            pass
        elif( '주식체결' in real_data_type):
            if( '종목번호' in item_dict):
                jongmok_code = item_dict['종목번호']
            if( '종목코드' in item_dict):
                jongmok_code = item_dict['종목코드']

            current_price  = 0
            if( '현재가' in item_dict):
                current_price = abs(int(item_dict['현재가']))

            # 실시간 데이터 대입 
            for index, col_name in enumerate(kw_util.dict_jusik[real_data_type]) :
                item_dict[col_name] = result_list[index]

        
    # 다음 codition list 를 감시 하기 위해 종목 섞기 
    def shuffleConditionOccurList(self):
        jongmok_info_dict = self.getConditionOccurList()

        if( jongmok_info_dict != None ):
            jongmok_code = jongmok_info_dict['종목코드']
            self.removeConditionOccurList(jongmok_code)
            self.conditionOccurList.append(jongmok_info_dict)

     # 실시간  주식 정보 요청 요청리스트 갱신  
     # WARNING: 실시간 요청도 TR 처럼 초당 횟수 제한이 있으므로 잘 사용해야함 
    def refreshRealRequest(self):
        # 버그로 모두 지우고 새로 등록하게 함 
        # print(util.whoami() )
        self.setRealRemove("ALL", "ALL")
        codeList  = []

        # 보유 잔고 실시간 정보 얻기 위해 추가 
        for code in self.jangoInfo.keys():
            if( code not in codeList):
                codeList.append(code)

        condition_list = self.getCodeListConditionOccurList()
        for code in condition_list: 
            if( code not in codeList):
                codeList.append(code)

        if( len(codeList) == 0 ):
            # 종목 미보유로 실시간 체결 요청 할게 없는 경우 코스닥 코스피 실시간 체결가가 올라오지 않으므로 임시로 하나 등록  
            codeList.append('044180')
        else:
            for code in codeList:
                if ( code not in user_setting.EXCEPTION_LIST):
                    self.addConditionOccurList(code)

        # 실시간 정보 요청 "0" 은 이전거 제외 하고 새로 요청
        if( len(codeList) ):
           #  WARNING: 주식 시세 실시간은 리턴되지 않음!
            # tmp = self.setRealReg(kw_util.sendRealRegSiseSrcNo, ';'.join(codeList), kw_util.type_fidset['주식시세'], "0")
            tmp = self.setRealReg(kw_util.sendRealRegHogaScrNo, ';'.join(codeList), kw_util.type_fidset['주식호가잔량'], "0")
            tmp = self.setRealReg(kw_util.sendRealRegChegyeolScrNo, ';'.join(codeList), kw_util.type_fidset['주식체결'], "0")
            tmp = self.setRealReg(kw_util.sendRealRegUpjongScrNo, '001;101', kw_util.type_fidset['업종지수'], "0")
            tmp = self.setRealReg(kw_util.sendRealRegTradeStartScrNo, '', kw_util.type_fidset['장시작시간'], "0")

    def make_excel(self, file_path, data_dict):
        # 주의 구글 스프레드 시트는 100개의 요청 제한이 있으므로  
        # 당일 정보만 한번에 batch_update 로 한번에 넣도록 함 
        wb = gc.open(user_setting.GOOGLE_SPREAD_SHEET_NAME)
        sheets = wb.worksheets()
        sheet_names = [ sheet.title for sheet in sheets]

        # 날짜별로 sheet name 저장  
        for date_str, value in sorted(data_dict.items()):
            # sheet name 이 존재 안하면 sheet add
            # sheet name 은 YYMM 형식 
            sheet_name = date_str
            new_sheet = None
            if( sheet_name not in sheet_names):
                new_sheet = wb.add_worksheet(title = sheet_name, rows = 100, cols = 20, index = 0 )
            else:
                continue

            rows = []
            stop_plus_range_list = []
            for row_index, line in enumerate(value):
                items = [ item.strip() for item in line.split('|') ]
                row_data = {}

                # 매도의 경우 정보 추가  
                if( len(items) == 11 ):
                    items.append( items[9])  # 종목이름
                    items.append( items[10].split(':')[0][8:])  # 매수시간
                    items.append( items[10].split(':')[1])  # 매수가격 
                    items.append( items[8])  # 매도시간 
                    items.append( items[6])  # 매도가격 
                range_str = "A{}:{}{}".format(row_index + 1, chr(ord('A') + len(items) ),  row_index + 1 )
                row_data['range'] = range_str                
                row_data['values'] = [ items ] 
                if( float(items[0]) > 0 ):
                    stop_plus_range_list.append(range_str)
                rows.append( row_data )
            new_sheet.batch_update(rows)
            for plus_range in stop_plus_range_list:
                new_sheet.format(plus_range, { "backgroundColor": { "red": 1.0, "green": 1.0, "blue": 0}})


        #  for openpyxl
        # result = os.path.isfile(file_path)

        # if( result == False):
        #     wb = openpyxl.Workbook()
        #     wb.save(file_path)

        # # excel open 
        # wb = openpyxl.load_workbook(file_path)

        # sheet_names = wb.sheetnames
        # # print(sheet_names)

        # # 날짜별로 sheet name 저장  
        # for date_str, value in sorted(data_dict.items()):
        #     # sheet name 이 존재 안하면 sheet add
        #     # sheet name 은 YYMM 형식 
        #     sheet_name = date_str
        #     new_sheet = None
        #     if( sheet_name not in sheet_names):
        #         new_sheet = wb.create_sheet(sheet_name)
        #     else:
        #         new_sheet = wb[sheet_name]

        #     for row_index, line in enumerate(value):
        #         items = [ item.strip() for item in line.split('|') ]

        #         # 매도의 경우 정보 추가  
        #         if( len(items) == 11 ):
        #             items.append( items[9])  # 종목이름
        #             items.append( items[10].split(':')[0][8:])  # 매수시간
        #             items.append( items[10].split(':')[1])  # 매수가격 
        #             items.append( items[8])  # 매도시간 
        #             items.append( items[6])  # 매도가격 

        #         for col_index, item in enumerate(items):
        #             new_sheet.cell(row = row_index + 1, column= col_index + 1 ).value = item

        # wb.save(file_path)
        print('excel save complete')

    # method 
    # 로그인
    # 0 - 성공, 음수값은 실패
    # 단순 API 호출이 되었느냐 안되었느냐만 확인 가능 
    @pyqtSlot(result=int)
    def commConnect(self):
        return self.ocx.dynamicCall("CommConnect()")

    # 로그인 상태 확인
    # 0:미연결, 1:연결완료, 그외는 에러
    @pyqtSlot(result=int)
    def getConnectState(self):
        return self.ocx.dynamicCall("GetConnectState()")

    # 로그 아웃
    @pyqtSlot()
    def commTerminate(self):
        self.ocx.dynamicCall("CommTerminate()")

    # 로그인한 사용자 정보를 반환한다.
    # “ACCOUNT_CNT” – 전체 계좌 개수를 반환한다.
    # "ACCNO" – 전체 계좌를 반환한다. 계좌별 구분은 ‘;’이다.
    # “USER_ID” - 사용자 ID를 반환한다.
    # “USER_NAME” – 사용자명을 반환한다.
    # “KEY_BSECGB” – 키보드보안 해지여부. 0:정상, 1:해지
    # “FIREW_SECGB” – 방화벽 설정 여부. 0:미설정, 1:설정, 2:해지
    @pyqtSlot(str, result=str)
    def getLoginInfo(self, tag):
        return self.ocx.dynamicCall("GetLoginInfo(QString)", [tag])

    # Tran 입력 값을 서버통신 전에 입력값일 저장한다.
    @pyqtSlot(str, str)
    def setInputValue(self, id, value):
        self.ocx.dynamicCall("SetInputValue(QString, QString)", id, value)


    @pyqtSlot(str, result=str)
    def getCodeListByMarket(self, sMarket):
        return self.ocx.dynamicCall("GetCodeListByMarket(QString)", sMarket)

    # 통신 데이터를 송신한다.
    # 0이면 정상
    # OP_ERR_SISE_OVERFLOW – 과도한 시세조회로 인한 통신불가
    # OP_ERR_RQ_STRUCT_FAIL – 입력 구조체 생성 실패
    # OP_ERR_RQ_STRING_FAIL – 요청전문 작성 실패
    # OP_ERR_NONE – 정상처리
    @pyqtSlot(str, str, int, str, result=int)
    def commRqData(self, rQName, trCode, prevNext, screenNo):
        return self.ocx.dynamicCall("CommRqData(QString, QString, int, QString)", rQName, trCode, prevNext, screenNo)

    # 수신 받은 데이터의 반복 개수를 반환한다.
    @pyqtSlot(str, str, result=int)
    def getRepeatCnt(self, trCode, recordName):
        return self.ocx.dynamicCall("GetRepeatCnt(QString, QString)", trCode, recordName)

    # Tran 데이터, 실시간 데이터, 체결잔고 데이터를 반환한다.
    # 1. Tran 데이터
    # 2. 실시간 데이터
    # 3. 체결 데이터
    # 1. Tran 데이터
    # sJongmokCode : Tran명
    # sRealType : 사용안함
    # sFieldName : 레코드명
    # nIndex : 반복인덱스
    # sInnerFieldName: 아이템명
    # 2. 실시간 데이터
    # sJongmokCode : Key Code
    # sRealType : Real Type
    # sFieldName : Item Index (FID)
    # nIndex : 사용안함
    # sInnerFieldName:사용안함
    # 3. 체결 데이터
    # sJongmokCode : 체결구분
    # sRealType : “-1”
    # sFieldName : 사용안함
    # nIndex : ItemIndex
    # sInnerFieldName:사용안함
    @pyqtSlot(str, str, str, int, str, result=str)
    def commGetData(self, jongmok_code, realType, fieldName, index, innerFieldName):
        return self.ocx.dynamicCall("CommGetData(QString, QString, QString, int, QString)", jongmok_code, realType, fieldName, index, innerFieldName).strip()

    # strRealType – 실시간 구분
    # nFid – 실시간 아이템
    # Ex) 현재가출력 - openApi.GetCommRealData(“주식시세”, 10);
    # 참고)실시간 현재가는 주식시세, 주식체결 등 다른 실시간타입(RealType)으로도 수신가능
    @pyqtSlot(str, int, result=str)
    def getCommRealData(self, realType, fid):
        return self.ocx.dynamicCall("GetCommRealData(QString, int)", realType, fid).strip()

    # 주식 주문을 서버로 전송한다.
    # sRQName - 사용자 구분 요청 명
    # sScreenNo - 화면번호[4]
    # sAccNo - 계좌번호[10]
    # nOrderType - 주문유형 (1:신규매수, 2:신규매도, 3:매수취소, 4:매도취소, 5:매수정정, 6:매도정정)
    # sCode, - 주식종목코드
    # nQty – 주문수량
    # nPrice – 주문단가
    # sHogaGb - 거래구분
    # sHogaGb – 00:지정가, 03:시장가, 05:조건부지정가, 06:최유리지정가, 07:최우선지정가, 10:지정가IOC, 13:시장가IOC, 16:최유리IOC, 20:지정가FOK, 23:시장가FOK, 26:최유리FOK, 61:장전시간외종가, 62:시간외단일가, 81:장후시간외종가
    # ※ 시장가, 최유리지정가, 최우선지정가, 시장가IOC, 최유리IOC, 시장가FOK, 최유리FOK, 장전시간외, 장후시간외 주문시 주문가격을 입력하지 않습니다.
    # ex)
    # 지정가 매수 - openApi.SendOrder(“RQ_1”, “0101”, “5015123410”, 1, “000660”, 10, 48500, “00”, “”);
    # 시장가 매수 - openApi.SendOrder(“RQ_1”, “0101”, “5015123410”, 1, “000660”, 10, 0, “03”, “”);
    # 매수 정정 - openApi.SendOrder(“RQ_1”,“0101”, “5015123410”, 5, “000660”, 10, 49500, “00”, “1”);
    # 매수 취소 - openApi.SendOrder(“RQ_1”, “0101”, “5015123410”, 3, “000660”, 10, “00”, “2”);
    # sOrgOrderNo – 원주문번호
    @pyqtSlot(str, str, str, int, str, int, int, str, str, result=int)
    def sendOrder(self, rQName, screenNo, accNo, orderType, code, qty, price, hogaGb, orgOrderNo):
        return self.ocx.dynamicCall("SendOrder(QString, QString, QString, int, QString, int, int, QString, QString)", [rQName, screenNo, accNo, orderType, code, qty, price, hogaGb, orgOrderNo])

    # 체결잔고 데이터를 반환한다.
    @pyqtSlot(int, result=str)
    def getChejanData(self, fid):
        return self.ocx.dynamicCall("GetChejanData(int)", fid)

    # 서버에 저장된 사용자 조건식을 가져온다.
    @pyqtSlot(result=int)
    def getConditionLoad(self):
        return self.ocx.dynamicCall("GetConditionLoad()")

    # 조건검색 조건명 리스트를 받아온다.
    # 조건명 리스트(인덱스^조건명)
    # 조건명 리스트를 구분(“;”)하여 받아온다
    @pyqtSlot(result=str)
    def getConditionNameList(self):
        return self.ocx.dynamicCall("GetConditionNameList()")

    # 조건검색 종목조회TR송신한다.
    # LPCTSTR strScrNo : 화면번호
    # LPCTSTR strConditionName : 조건명
    # int nIndex : 조건명인덱스
    # int nSearch : 조회구분(0:일반조회, 1:실시간조회, 2:연속조회)
    # 1:실시간조회의 화면 개수의 최대는 10개
    @pyqtSlot(str, str, int, int)
    def sendCondition(self, scrNo, conditionName, index, search):
        self.ocx.dynamicCall("SendCondition(QString,QString, int, int)", scrNo, conditionName, index, search)

    # 실시간 조건검색을 중지합니다.
    # ※ 화면당 실시간 조건검색은 최대 10개로 제한되어 있어서 더 이상 실시간 조건검색을 원하지 않는 조건은 중지해야만 카운트 되지 않습니다.
    @pyqtSlot(str, str, int)
    def sendConditionStop(self, scrNo, conditionName, index):
        self.ocx.dynamicCall("SendConditionStop(QString, QString, int)", scrNo, conditionName, index)

    # 복수종목조회 Tran을 서버로 송신한다.
    # OP_ERR_RQ_STRING – 요청 전문 작성 실패
    # OP_ERR_NONE - 정상처리
    #
    # sArrCode – 종목간 구분은 ‘;’이다.
    # nTypeFlag – 0:주식관심종목정보, 3:선물옵션관심종목정보
    @pyqtSlot(str, bool, int, int, str, str)
    def commKwRqData(self, arrCode, next, codeCount, typeFlag, rQName, screenNo):
    	self.ocx.dynamicCall("CommKwRqData(QString, QBoolean, int, int, QString, QString)", arrCode, next, codeCount, typeFlag, rQName, screenNo)

    # 실시간 등록을 한다.
    # strScreenNo : 화면번호
    # strCodeList : 종목코드리스트(ex: 039490;005930;…)
    # strFidList : FID번호(ex:9001;10;13;…)
    # 	9001 – 종목코드
    # 	10 - 현재가
    # 	13 - 누적거래량
    # strOptType : 타입(“0”, “1”)
    # 타입 “0”은 항상 마지막에 등록한 종목들만 실시간등록이 됩니다.
    # 타입 “1”은 이전에 실시간 등록한 종목들과 함께 실시간을 받고 싶은 종목을 추가로 등록할 때 사용합니다.
    # ※ 종목, FID는 각각 한번에 실시간 등록 할 수 있는 개수는 100개 입니다.
    @pyqtSlot(str, str, str, str,  result=int)
    def setRealReg(self, screenNo, codeList, fidList, optType):
        return self.ocx.dynamicCall("SetRealReg(QString, QString, QString, QString)", screenNo, codeList, fidList, optType)

    # 종목별 실시간 해제
    # strScrNo : 화면번호
    # strDelCode : 실시간 해제할 종목코드
    # -화면별 실시간해제
    # 여러 화면번호로 걸린 실시간을 해제하려면 파라메터의 화면번호와 종목코드에 “ALL”로 입력하여 호출하시면 됩니다.
    # SetRealRemove(“ALL”, “ALL”);
    # 개별화면별로 실시간 해제 하시려면 파라메터에서 화면번호는 실시간해제할
    # 화면번호와 종목코드에는 “ALL”로 해주시면 됩니다.
    # SetRealRemove(“0001”, “ALL”);
    # -화면의 종목별 실시간해제
    # 화면의 종목별로 실시간 해제하려면 파라메터에 해당화면번호와 해제할
    # 종목코드를 입력하시면 됩니다.
    # SetRealRemove(“0001”, “039490”);
    # 문서와는 달리 return 없음 
    # SetRealReg 로 등록한 함수만 해제 가능 
    @pyqtSlot(str, str)
    def setRealRemove(self, scrNo, delCode):
        self.ocx.dynamicCall("SetRealRemove(QString, QString)", scrNo, delCode)
        
        
    # 수신 데이터를 반환한다. 
    # LPCTSTR strTrCode : 조회한TR코드
    # LPCTSTR strRecordName: 조회한 TR명
    # nIndex : 복수 데이터 인덱스
    # strItemName: 아이템 명
    # 반환값: 수신 데이터
    
    @pyqtSlot(str, str, int, str, result=str)
    def getCommData(self, trCode, recordName, index, itemName):
        return self.ocx.dynamicCall("GetCommData(QString, QString, int, QString)", 
        trCode, recordName, index, itemName)

    # 차트 조회한 데이터 전부를 배열로 받아온다.
    # LPCTSTR strTrCode : 조회한TR코드
    # LPCTSTR strRecordName: 조회한 TR명
    # ※항목의 위치는 KOA Studio의 TR목록 순서로 데이터를 가져옵니다.
    # 예로 OPT10080을 살펴보면 OUTPUT의 멀티데이터의 항목처럼 현재가, 거래량, 체결시간등 순으로 항목의 위치가 0부터 1씩증가합니다.
    @pyqtSlot(str, str, result=str)
    def getCommDataEx(self, trCode, recordName):
        return self.ocx.dynamicCall("GetCommDataEx(QString, QString)", trCode, recordName)

    # 리얼 시세를 끊는다.
    # 화면 내 모든 리얼데이터 요청을 제거한다.
    # 화면을 종료할 때 반드시 위 함수를 호출해야 한다.
    # Ex) openApi.DisconnectRealData(“0101”);
    @pyqtSlot(str)
    def disconnectRealData(self, scnNo):
        self.ocx.dynamicCall("DisconnectRealData(QString)", scnNo)

    # 종목코드의 한글명을 반환한다.
    # strCode – 종목코드
    # 종목한글명
    # 없는 코드 일경우 empty 를 리턴함 
    @pyqtSlot(str, result=str)
    def getMasterCodeName(self, strCode):
        return self.ocx.dynamicCall("GetMasterCodeName(QString)", strCode)

    # 설명 종목코드의 전일가를 반환한다. 
    # 입력값: strCode – 종목코드 
    # 반환값: 전일가  
    @pyqtSlot(str, result=str)
    def GetMasterLastPrice(self, strCode):
        return self.ocx.dynamicCall("GetMasterLastPrice(QString)", strCode)

    # 종목코드의 한다.
    # strCode – 종목코드
    # 입력한 종목에 대한 대분류, 중분류, 업종구분값을 구분자로 연결한 문자열을 얻을수 있습니다.(여기서 구분자는 '|'와 ';'입니다.) 
    # KOA_Functions("GetMasterStockInfo", 종목코드) 
    # 시장구분0|코스닥|벤처기업;시장구분1|소형주;업종구분|제조|기계/장비
    # 시장구분0|거래소;시장구분1|중형주;업종구분|서비스업;
    @pyqtSlot(str, result=str)
    def getMasterStockInfo(self, strCode):
        stock_info = self.ocx.dynamicCall("KOA_Functions(QString, QString)", "GetMasterStockInfo", strCode)
        # api return 버그로 추가 해줌 
        kospi_kosdaq = ''
        yupjong = ''

        if( stock_info != ''):
            if( stock_info[-1] == ';'):
                stock_info = stock_info[0:-1]
            kospi_kosdaq = stock_info.split(';')[0].split('|')[1]
            yupjong = stock_info.split(';')[-1].split('|')[-1]
        return kospi_kosdaq + ':' + yupjong

if __name__ == "__main__":
    def test_buy():
        # 비정상 매수 (시장가에 단가 넣기 ) 우리종금 1주  
        # objKiwoom.sendOrder("buy", kw_util.sendOrderScreenNo, objKiwoom.account_list[0], kw_util.dict_order["신규매수"], 
        # "010050", 1, 900 , kw_util.dict_order["시장가"], "")

        # 정상 매수 - kd 건설 1주 
        objKiwoom.sendOrder("buy", kw_util.sendOrderScreenNo, objKiwoom.account_list[0], kw_util.dict_order["신규매수"], 
        "044180", 1, 0 , kw_util.dict_order["시장가"], "")
        pass

    def test_sell():
        #정상 매도 - kd 건설 1주 
        objKiwoom.sendOrder("sell", kw_util.sendOrderScreenNo, objKiwoom.account_list[0], kw_util.dict_order["신규매도"], 
        "044180", 1, 0 , kw_util.dict_order["시장가"], "")
        pass

    def test_condition():
        objKiwoom._OnReceiveRealCondition("044180", "I",  "단타 추세", 1)
        pass
    def test_getCodeList():
        # 코스피 코드 리스트
        result = objKiwoom.getCodeListByMarket('0')
        print(result)
        # 코스닥 코드 리스트 
        result = objKiwoom.getCodeListByMarket('10')
        print(result)
        pass
    def test_jusikGibon():
        objKiwoom.requestOpt10001("044180")
        pass
    def test_jusik_condition_occur():
        objKiwoom.addConditionOccurList('044180')
        pass

    @pyqtSlot()
    def test_make_jangoInfo():
        objKiwoom.makeJangoInfoFile()
        pass
    def test_make_chegyeolInfo():
        objKiwoom.makeChegyeolInfoFile()
    def test_terminate():
        objKiwoom.sigTerminating.emit()
        pass
    
    def test_make_excel():
        file_path = r'd:\download\통합 문서2.xlsx'
        test_data = {  
            "170822": [
                "      8.56|      8750| 매수횟수: 1 | (익절이다) | 210540 | -매도 |      22200 |       5 | 141905 | 디와이파워 ",
                "         0|         0| 매수횟수: 1 | (단순매수) | 038870 | +매수 |      10350 |      10 | 142205 | 에코바이오 "
            ],
            "170824": [
                "      8.81|      9180| 매수횟수: 1 | (익절이다) | 033100 | -매도 |       6300 |      18 | 090136 | 제룡전기 ",
                "         0|         0| 매수횟수: 1 | (단순매수) | 069540 | +매수 |       6200 |      17 | 090349 | 라이트론 ",
                "      8.55|      9010| 매수횟수: 1 | (익절미달) | 069540 | -매도 |       6730 |      17 | 090557 | 라이트론 ",
                "      8.76|      8925| 매수횟수: 1 | (익절미달) | 226350 | -매도 |       3165 |      35 | 090842 | 아이엠텍 ",
                "         0|         0| 매수횟수: 1 | (단순매수) | 180400 | +매수 |      14250 |       8 | 091259 | 엠지메드 ",
                "         0|         0| 매수횟수: 1 | (단순매수) | 072470 | +매수 |       8247 |      13 | 091845 | 우리산업홀딩스 ",
                "         0|         0| 매수횟수: 4 | (단순매수) | 031860 | +매수 |       5050 |      80 | 094752 | 엔에스엔 "
            ],
            "170914": [
                "     13.53|     14000| 매수횟수: 1 | (익절이다) | 038870 | -매도 |      11750 |      10 | 090102 | 에코바이오 ",
                "         0|         0| 매수횟수: 1 | (단순매수) | 023800 | +매수 |       6400 |       8 | 090105 | 인지컨트롤스 ",
                "         0|         0| 매수횟수: 1 | (단순매수) | 042600 | +매수 |       8420 |       6 | 090113 | 새로닉스 ",
                "      8.99|      4720| 매수횟수: 1 | (익절미달) | 161570 | -매도 |       7150 |       8 | 093611 | 미동앤씨네마 ",
                "         0|         0| 매수횟수: 1 | (단순매수) | 005420 | +매수 |      19050 |       3 | 093933 | 코스모화학 "
            ]
        }
        objKiwoom.make_excel(file_path, test_data)
    def test_business_day():
        base_time = datetime.datetime.strptime('20190920090011', '%Y%m%d%H%M%S')

        target_day = util.date_by_adding_business_days(base_time, 1 )
        print(target_day)


    # putenv 는 current process 에 영향을 못끼치므로 environ 에서 직접 세팅 
    # qml debugging 를 위해 QML_IMPORT_TRACE 환경변수 1로 세팅 후 DebugView 에서 디버깅 메시지 확인 가능  
    os.environ['QML_IMPORT_TRACE'] = '1'
    # print(os.environ['QML_IMPORT_TRACE'])
    myApp = QtWidgets.QApplication(sys.argv)
    objKiwoom = KiwoomConditon()

    form = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(form)

    event_filter = CloseEventEater()
    form.installEventFilter( event_filter )

    ui.btnMakeExcel.clicked.connect(objKiwoom.onBtnMakeExcelClicked )
    ui.btnStart.clicked.connect(objKiwoom.onBtnStartClicked)

    ui.btnYupjong.clicked.connect(objKiwoom.onBtnYupjongClicked)
    ui.btnJango.clicked.connect(objKiwoom.onBtnJangoClicked)
    ui.btnChegyeol.clicked.connect(objKiwoom.onBtnChegyeolClicked)
    ui.btnCondition.clicked.connect(objKiwoom.onBtnConditionClicked)

    ui.lineCmd.textChanged.connect(objKiwoom.onLineCmdTextChanged)
    ui.btnRun.clicked.connect(objKiwoom.onBtnRunClicked)

    form.show()
    # test_business_day()

    logging.basicConfig(filename='system_err.log', filemode='a',format='%(asctime)s - %(message)s', level=logging.INFO)

    # try:
    #     1/0 
    # except Exception as e:
    #     logging.exception("Error Occured")

    sys.exit(myApp.exec_())